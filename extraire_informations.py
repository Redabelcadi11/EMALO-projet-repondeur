from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz
from src.runtime_paths import (
    bootstrap_runtime_environment,
    get_project_root,
)
from src.clients import (
    charger_aliases_telephoniques_confirmes,
    charger_telephones_clients,
    charger_variantes_clients,
    candidats_pour_arbitrage_llm,
    client_requiert_arbitrage_llm,
    enrichir_alias_avec_variantes,
    enrichir_clients_avec_aliases_telephoniques_confirmes,
    enrichir_clients_avec_telephones,
    filtrer_mentions_client_resolu,
    identifier_client,
    normaliser_telephone,
    normaliser_telephones,
)
from src.produits import (
    _preuve_positive_noyau_produit,
    charger_synonymes_produits,
    construire_catalogue_global,
    decouper_clauses_produits as decouper_clauses_produits_v2,
    extraire_mentions_produits as extraire_mentions_produits_v2,
    chercher_produits as chercher_produits_v2,
)

bootstrap_runtime_environment()

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------

RACINE_PROJET = get_project_root()

DOSSIER_TRANSCRIPTIONS = RACINE_PROJET / "resultats" / "transcriptions"
DOSSIER_RESULTATS = RACINE_PROJET / "resultats" / "extractions"
DOSSIER_CONFIG = RACINE_PROJET / "config"
DOSSIER_COMMANDES_VALIDEES = RACINE_PROJET / "resultats" / "commandes-validees"
DOSSIER_COMMANDES_PROBLEMATIQUES = RACINE_PROJET / "resultats" / "commandes-problematiques"

DOSSIER_CLIENTS = RACINE_PROJET / "ressources-originales" / "informations-clients"
DOSSIER_CADENCIER = RACINE_PROJET / "ressources-originales" / "cadencier"
CHEMIN_CADENCIER_COMPLEMENT = (
    DOSSIER_CADENCIER
    / "historique_es_pretest_2026-05-20_au_2026-06-22.csv"
)
CHEMIN_VARIANTES_CLIENTS = DOSSIER_CONFIG / "variantes-clients.json"
CHEMIN_TELEPHONES_CLIENTS = DOSSIER_CONFIG / "telephones-clients.json"
CHEMIN_ALIASES_TELEPHONIQUES_CONFIRMES = (
    DOSSIER_CONFIG / "aliases-telephoniques-confirmes.json"
)
CHEMIN_COMMANDES_COURTES_CLIENTS = (
    DOSSIER_CONFIG / "commandes-courtes-clients.json"
)
CHEMIN_SYNONYMES_PRODUITS = DOSSIER_CONFIG / "synonymes-produits.json"
CHEMIN_UNITES_ARTICLES = DOSSIER_CONFIG / "unites-articles.csv"
CHEMIN_CATALOGUE_ARTICLES = DOSSIER_CONFIG / "catalogue-articles.json"
CHEMIN_REFERENCES_CONTROLE = (
    DOSSIER_CONFIG / "references-articles-controle.json"
)

NOMBRES = {
    "un": 1,
    "une": 1,
    "deux": 2,
    "trois": 3,
    "quatre": 4,
    "cinq": 5,
    "six": 6,
    "sept": 7,
    "huit": 8,
    "neuf": 9,
    "dix": 10,
    "onze": 11,
    "douze": 12,
    "treize": 13,
    "quatorze": 14,
    "quinze": 15,
    "seize": 16,
    "vingt": 20,
}

UNITES = {
    "kg": "KG",
    "kilo": "KG",
    "kilos": "KG",
    "kilogramme": "KG",
    "kilogrammes": "KG",
    "gramme": "G",
    "grammes": "G",
    "litre": "L",
    "litres": "L",
    "boite": "BOITE",
    "boites": "BOITE",
    "carton": "CAR",
    "cartons": "CAR",
    "colis": "COL",
    "piece": "PCE",
    "pieces": "PCE",
    "palette": "PAL",
    "palettes": "PAL",
}

JOURS_SEMAINE = {
    "lundi": 0,
    "mardi": 1,
    "mercredi": 2,
    "jeudi": 3,
    "vendredi": 4,
    "samedi": 5,
    "dimanche": 6,
}

MOTIFS_SUPPRESSION_COMMANDE = (
    "supprimer",
    "supprime",
    "retirer",
    "retire",
    "enlever",
    "enleve",
    "annuler",
    "annule",
)

MOTIFS_NOUVELLE_COMMANDE = (
    "ajout",
    "ajouter",
    "ajoute",
    "complement",
    "modifier",
    "modifie",
    "modification",
    "changer",
    "change",
    "remplacer",
    "remplace",
)

MOTIFS_MODIFICATION_A_RAPPELER = (
    r"\bremplac\w*\b.+\bpar\b",
    r"\b(?:a|en)\s+la\s+place\b",
    r"\bne\s+mett\w*\s+pas\b",
    r"\b(?:dans|sur)\s+(?:ma|la|notre)\s+commande\b.+\bpas\s+de\b",
    r"\b(?:j\s+en\s+ai|je\s+viens\s+de)\s+retrouv\w*\b",
    r"\bannulez\b",
    r"\bchang\w*\b.+\bquantit\w*\b",
)

MOTIFS_RAPPEL_CLIENT = (
    "rappelle moi",
    "rappelez moi",
    "me rappeler",
    "vous me rappeliez",
    "appelez moi",
    "me rappeleriez",
)

MOTS_GENERIQUES_CLIENT = {
    "le",
    "la",
    "les",
    "un",
    "une",
    "restaurant",
    "resto",
    "hotel",
    "bar",
    "bistrot",
    "brasserie",
    "sarl",
    "sas",
    "sasu",
    "eurl",
    "societe",
}


# -------------------------------------------------------------------
# Normalisation
# -------------------------------------------------------------------

def enlever_accents(texte: str) -> str:
    texte = unicodedata.normalize("NFD", texte)

    return "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )


def normaliser(valeur: Any) -> str:
    texte = "" if valeur is None else str(valeur)

    texte = enlever_accents(texte).lower()

    texte = re.sub(
        r"[^a-z0-9]+",
        " ",
        texte,
    )

    return re.sub(
        r"\s+",
        " ",
        texte,
    ).strip()


def normaliser_nom_colonne(valeur: Any) -> str:
    """Normalise un en-tete sans tenir compte de ses separateurs."""

    return re.sub(
        r"[^a-z0-9]+",
        "",
        enlever_accents(str(valeur or "")).lower(),
    )


def simplifier_nom_client(texte: str) -> str:
    tokens = normaliser(texte).split()

    tokens = [
        token
        for token in tokens
        if token not in MOTS_GENERIQUES_CLIENT
    ]

    return " ".join(tokens)


def creer_alias_client(nom_client: str) -> list[str]:
    """
    Exemple :
    LE BISTROT DES FILLES SARL TUNA

    devient notamment :
    - le bistrot des filles sarl tuna
    - le bistrot des filles
    - filles
    """

    nom_normalise = normaliser(nom_client)

    partie_avant_statut = re.split(
        r"\b(?:sarl|sas|sasu|eurl)\b",
        nom_normalise,
        maxsplit=1,
    )[0].strip()

    aliases = {
        nom_normalise,
        partie_avant_statut,
        simplifier_nom_client(nom_normalise),
        simplifier_nom_client(partie_avant_statut),
    }

    for morceau in re.split(r"\s+-\s+|/|\(|\)", str(nom_client or "")):
        morceau_normalise = normaliser(morceau)

        if not morceau_normalise:
            continue

        morceau_avant_statut = re.split(
            r"\b(?:sarl|sas|sasu|eurl)\b",
            morceau_normalise,
            maxsplit=1,
        )[0].strip()

        if morceau_avant_statut:
            aliases.add(morceau_avant_statut)
            aliases.add(
                simplifier_nom_client(morceau_avant_statut)
            )

    return sorted(
        alias
        for alias in aliases
        if alias
    )


# -------------------------------------------------------------------
# Lecture des fichiers Excel
# -------------------------------------------------------------------

def lister_fichiers_xlsx_metier(dossier: Path) -> list[Path]:
    """Retourne les classeurs reels, sans les verrous temporaires d'Excel."""

    return sorted(
        chemin
        for chemin in dossier.glob("*.xlsx")
        if chemin.is_file() and not chemin.name.startswith("~$")
    )


def trouver_fichier_xlsx(dossier: Path) -> Path:
    fichiers = lister_fichiers_xlsx_metier(dossier)

    if not fichiers:
        raise FileNotFoundError(
            f"Aucun fichier .xlsx trouvé dans : {dossier}"
        )

    return fichiers[0]


def trouver_fichiers_xlsx(dossier: Path) -> list[Path]:
    fichiers = lister_fichiers_xlsx_metier(dossier)

    if not fichiers:
        raise FileNotFoundError(
            f"Aucun fichier .xlsx trouvÃ© dans : {dossier}"
        )

    return fichiers


def lire_lignes_xlsx(
    chemin: Path,
) -> tuple[list[str], list[dict[str, Any]]]:
    classeur = load_workbook(
        chemin,
        read_only=True,
        data_only=True,
    )

    feuille = classeur.active

    lignes_excel = feuille.iter_rows(values_only=True)

    en_tetes = [
        str(valeur or "").strip()
        for valeur in next(lignes_excel)
    ]

    lignes: list[dict[str, Any]] = []

    for valeurs in lignes_excel:
        ligne = {
            en_tetes[index]: valeurs[index]
            for index in range(
                min(len(en_tetes), len(valeurs))
            )
        }

        if any(
            valeur not in (None, "")
            for valeur in ligne.values()
        ):
            lignes.append(ligne)

    classeur.close()

    return en_tetes, lignes


def choisir_colonne(
    en_tetes: list[str],
    candidats: list[str],
    obligatoire: bool = True,
    *,
    champ_metier: str = "colonne",
    chemin: Path | None = None,
) -> str | None:
    candidats_normalises = [
        normaliser_nom_colonne(candidat)
        for candidat in candidats
    ]
    en_tetes_normalises = {
        en_tete: normaliser_nom_colonne(en_tete)
        for en_tete in en_tetes
    }

    for candidat in candidats_normalises:
        for en_tete, en_tete_normalisee in en_tetes_normalises.items():
            if candidat == en_tete_normalisee:
                return en_tete

    # Certains exports ajoutent une periode au nom d'une mesure. Une fois les
    # aliases exacts testes, le prefixe reconnait ce suffixe variable sans
    # confondre une courte sous-chaine avec un autre champ metier.
    for candidat in candidats_normalises:
        for en_tete, en_tete_normalisee in en_tetes_normalises.items():
            if candidat and en_tete_normalisee.startswith(candidat):
                return en_tete

    if obligatoire:
        fichier = str(chemin) if chemin is not None else "<fichier non precise>"
        raise KeyError(
            "Colonne obligatoire introuvable.\n"
            f"Champ metier recherche : {champ_metier}\n"
            f"Aliases essayes : {candidats}\n"
            f"Colonnes reellement disponibles : {en_tetes}\n"
            f"Fichier concerne : {fichier}"
        )

    return None


# -------------------------------------------------------------------
# Clients
# -------------------------------------------------------------------

ALIASES_COLONNES_CLIENTS: dict[str, list[str]] = {
    "code_client": [
        "clients code",
        "client code",
        "clients livres code",
        "client livre code",
        "code client",
        "n cpte",
        "numero compte",
        "compte",
    ],
    "libelle_client": [
        "clients lib",
        "client lib",
        "clients livres lib",
        "client livre lib",
        "nom client",
        "raison sociale",
    ],
    "ville": ["ville"],
    "adresse_1": ["adresse 1", "adresse1"],
    "adresse_2": ["adresse 2", "adresse2"],
    "code_postal": ["code postal", "codepostal", "cp"],
    "telephone": [
        "telephone du contact",
        "telephone contact",
        "telephone",
        "tel",
        "portable",
    ],
    "code_recherche": ["code recherche", "cod rech"],
}

def charger_clients() -> list[dict[str, Any]]:
    clients_par_code: dict[str, dict[str, Any]] = {}

    for chemin in trouver_fichiers_xlsx(DOSSIER_CLIENTS):
        en_tetes, lignes = lire_lignes_xlsx(chemin)

        col_code = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["code_client"],
            champ_metier="code_client",
            chemin=chemin,
        )

        col_nom = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["libelle_client"],
            champ_metier="libelle_client",
            chemin=chemin,
        )

        col_ville = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["ville"],
            obligatoire=False,
            champ_metier="ville",
            chemin=chemin,
        )
        col_adresse_1 = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["adresse_1"],
            obligatoire=False,
            champ_metier="adresse_1",
            chemin=chemin,
        )
        col_adresse_2 = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["adresse_2"],
            obligatoire=False,
            champ_metier="adresse_2",
            chemin=chemin,
        )
        col_code_postal = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["code_postal"],
            obligatoire=False,
            champ_metier="code_postal",
            chemin=chemin,
        )
        col_telephone = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["telephone"],
            obligatoire=False,
            champ_metier="telephone",
            chemin=chemin,
        )
        col_code_recherche = choisir_colonne(
            en_tetes,
            ALIASES_COLONNES_CLIENTS["code_recherche"],
            obligatoire=False,
            champ_metier="code_recherche",
            chemin=chemin,
        )

        for ligne in lignes:
            code = str(ligne.get(col_code, "") or "").strip()
            nom = str(ligne.get(col_nom, "") or "").strip()

            if not code or not nom:
                continue

            ville = (
                str(ligne.get(col_ville, "") or "").strip()
                if col_ville
                else ""
            )
            adresse_1 = (
                str(ligne.get(col_adresse_1, "") or "").strip()
                if col_adresse_1
                else ""
            )
            adresse_2 = (
                str(ligne.get(col_adresse_2, "") or "").strip()
                if col_adresse_2
                else ""
            )
            code_postal = (
                str(ligne.get(col_code_postal, "") or "").strip()
                if col_code_postal
                else ""
            )
            telephone = (
                str(ligne.get(col_telephone, "") or "").strip()
                if col_telephone
                else ""
            )
            telephones = normaliser_telephones(telephone)

            aliases = set(
                creer_alias_client(nom)
            )

            aliases.add(
                normaliser(code)
            )
            if col_code_recherche:
                aliases.add(
                    normaliser(ligne.get(col_code_recherche, ""))
                )

            nouveau_client = {
                "code_client": code,
                "nom_client": nom,
                "ville": ville,
                "adresse_1": adresse_1,
                "adresse_2": adresse_2,
                "code_postal": code_postal,
                "telephone": telephone,
                "telephones": telephones,
                "telephones_info": list(telephones),
                "aliases": sorted(alias for alias in aliases if alias),
            }

            existant = clients_par_code.get(code)
            if not existant:
                clients_par_code[code] = nouveau_client
                continue

            for cle in [
                "nom_client",
                "ville",
                "adresse_1",
                "adresse_2",
                "code_postal",
                "telephone",
            ]:
                if (
                    not str(existant.get(cle, "") or "").strip()
                    and str(nouveau_client.get(cle, "") or "").strip()
                ):
                    existant[cle] = nouveau_client[cle]

            existant["telephones"] = sorted(
                {
                    *[
                        str(value)
                        for value in existant.get("telephones", [])
                        if str(value).strip()
                    ],
                    *[
                        str(value)
                        for value in nouveau_client.get("telephones", [])
                        if str(value).strip()
                    ],
                }
            )
            existant["telephones_info"] = sorted(
                {
                    *[
                        str(value)
                        for value in existant.get("telephones_info", [])
                        if str(value).strip()
                    ],
                    *[
                        str(value)
                        for value in nouveau_client.get("telephones_info", [])
                        if str(value).strip()
                    ],
                }
            )
            existant["aliases"] = sorted(
                {
                    *[
                        str(value)
                        for value in existant.get("aliases", [])
                        if str(value).strip()
                    ],
                    *[
                        str(value)
                        for value in nouveau_client.get("aliases", [])
                        if str(value).strip()
                    ],
                }
            )

    return list(clients_par_code.values())


def extraire_zone_presentation_client(
    transcription: str,
) -> str:
    """
    On cherche le client uniquement au début du message.

    Exemple :
    Bonjour, c'est Les Affranchis, je voudrais...
    """

    texte = normaliser(transcription)

    marqueurs_fin = [
        "je voudrais",
        "je souhaite",
        "je souhaiterais",
        "je commande",
        "il me faudrait",
        "pour demain",
    ]

    positions = [
        texte.find(marqueur)
        for marqueur in marqueurs_fin
        if texte.find(marqueur) != -1
    ]

    if positions:
        texte = texte[:min(positions)]

    return texte.strip()


def calculer_score_client(
    zone_client: str,
    client: dict[str, Any],
) -> float:
    zone_normale = normaliser(zone_client)
    zone_simplifiee = simplifier_nom_client(zone_client)

    meilleur_score = 0.0

    for alias in client["aliases"]:
        alias_normalise = normaliser(alias)
        alias_simplifie = simplifier_nom_client(alias)

        if not alias_normalise:
            continue

        # Correspondance exacte du code ou du nom dans la présentation.
        if (
            len(alias_normalise) >= 3
            and alias_normalise in zone_normale
        ):
            meilleur_score = max(meilleur_score, 100.0)

        if (
            len(alias_simplifie) >= 3
            and alias_simplifie in zone_simplifiee
        ):
            meilleur_score = max(meilleur_score, 100.0)

        if len(alias_simplifie) >= 4:
            meilleur_score = max(
                meilleur_score,
                float(
                    fuzz.token_set_ratio(
                        alias_simplifie,
                        zone_simplifiee,
                    )
                ),
                float(
                    fuzz.partial_ratio(
                        alias_simplifie,
                        zone_simplifiee,
                    )
                ),
            )

    return round(meilleur_score, 2)


def chercher_clients(
    transcription: str,
    clients: list[dict[str, Any]],
    limite: int = 5,
) -> list[dict[str, Any]]:
    zone_client = extraire_zone_presentation_client(
        transcription
    )

    candidats: list[dict[str, Any]] = []

    for client in clients:
        score = calculer_score_client(
            zone_client,
            client,
        )

        if score >= 45:
            candidats.append(
                {
                    "code_client": client["code_client"],
                    "nom_client": client["nom_client"],
                    "ville": client["ville"],
                    "score": score,
                }
            )

    candidats.sort(
        key=lambda candidat: candidat["score"],
        reverse=True,
    )

    return candidats[:limite]


def client_est_fiable(
    candidats: list[dict[str, Any]],
) -> bool:
    if not candidats:
        return False

    premier = candidats[0]["score"]

    deuxieme = (
        candidats[1]["score"]
        if len(candidats) > 1
        else 0
    )

    return (
        premier >= 88
        and (
            premier - deuxieme >= 5
            or premier == 100
        )
    )


# -------------------------------------------------------------------
# Cadencier
# -------------------------------------------------------------------

@lru_cache(maxsize=1)
def _charger_index_references_officielles(
) -> tuple[set[str], dict[str, str], dict[str, str]]:
    """Indexe les codes actuels et les libelles officiels non ambigus."""
    if not CHEMIN_REFERENCES_CONTROLE.exists():
        return set(), {}, {}
    try:
        payload = json.loads(
            CHEMIN_REFERENCES_CONTROLE.read_text(encoding="utf-8")
        )
    except (OSError, json.JSONDecodeError):
        return set(), {}, {}
    references = (
        payload.get("references", {})
        if isinstance(payload, dict)
        else {}
    )
    if not isinstance(references, dict):
        return set(), {}, {}

    codes = {
        str(code).strip()
        for code in references
        if str(code).strip()
    }
    codes_reels: set[str] = set()
    references_par_libelle: dict[
        str,
        list[tuple[str, str]],
    ] = defaultdict(list)
    for code_brut, reference in references.items():
        code = str(code_brut).strip()
        if not code or not isinstance(reference, dict):
            continue
        suffixe_duplique = re.match(r"^(.+)\.(\d+)$", code)
        if suffixe_duplique and suffixe_duplique.group(1) in codes:
            # Suffixe technique cree par l'importeur pour une seconde ligne
            # de langue du meme code source. Ce n'est pas un code ERP.
            continue
        codes_reels.add(code)
        libelle_brut = str(reference.get("label") or "").strip()
        libelle = normaliser(libelle_brut)
        if libelle:
            references_par_libelle[libelle].append(
                (code, libelle_brut)
            )

    libelles_uniques = {
        libelle: references_libelle[0][0]
        for libelle, references_libelle in references_par_libelle.items()
        if len({code for code, _ in references_libelle}) == 1
    }
    alias_par_code: dict[str, str] = {}
    for references_libelle in references_par_libelle.values():
        codes_libelle = {code for code, _ in references_libelle}
        if len(codes_libelle) <= 1:
            continue
        codes_non_marques = {
            code
            for code, libelle_brut in references_libelle
            if "*" not in libelle_brut
        }
        if len(codes_non_marques) != 1:
            continue
        code_canonique = next(iter(codes_non_marques))
        for code, libelle_brut in references_libelle:
            if code != code_canonique and "*" in libelle_brut:
                alias_par_code[code] = code_canonique
    return codes_reels, libelles_uniques, alias_par_code


def _canonicaliser_code_article_reference(
    code_article: str,
    libelle_article: str,
) -> str:
    """Remplace un ancien code par le code officiel du meme libelle exact."""
    code = str(code_article or "").strip()
    if not code:
        return ""
    codes_officiels, codes_par_libelle, alias_par_code = (
        _charger_index_references_officielles()
    )
    if code in alias_par_code:
        return alias_par_code[code]
    if code in codes_officiels:
        return code
    libelle = normaliser(str(libelle_article or ""))
    return codes_par_libelle.get(libelle, code)


def _fusionner_historique_cadencier_pretest(
    produits_par_client: dict[str, dict[str, dict[str, Any]]],
    chemin_csv: Path = CHEMIN_CADENCIER_COMPLEMENT,
) -> None:
    """Ajoute un historique anterieur au corpus avec un garde-fou anti-fuite."""
    if not chemin_csv.exists():
        return

    chemin_manifest = chemin_csv.with_suffix(".manifest.json")
    if not chemin_manifest.exists():
        raise RuntimeError(
            f"Manifest du complement cadencier introuvable: {chemin_manifest}"
        )

    manifest = json.loads(chemin_manifest.read_text(encoding="utf-8"))
    if manifest.get("truth_from_evaluation_included") is not False:
        raise RuntimeError("Complement cadencier refuse: verite evaluation non isolee")
    date_coupure = date.fromisoformat(str(manifest["cutoff_inclusive"]))
    date_premier_audio = date.fromisoformat(
        str(manifest["evaluation_audio_from"])
    )
    if date_coupure >= date_premier_audio:
        raise RuntimeError(
            "Complement cadencier refuse: coupure posterieure au debut du corpus"
        )

    seuil_recent = date_coupure - timedelta(days=30)
    for produits in produits_par_client.values():
        for produit in produits.values():
            produit["nb_ventes_article_recentes"] = 0

    historiques_quantites: dict[
        tuple[str, str], Counter[float]
    ] = defaultdict(Counter)
    recence_quantites: dict[tuple[str, str], dict[float, int]] = defaultdict(dict)

    with chemin_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        for ligne in csv.DictReader(handle, delimiter=";"):
            code_client = str(ligne.get("client_code") or "").strip()
            code_article = str(ligne.get("article_code") or "").strip()
            libelle = str(ligne.get("designation") or "").strip()
            if not code_client or not code_article or not libelle:
                continue
            code_article = _canonicaliser_code_article_reference(
                code_article,
                libelle,
            )

            texte_date = str(
                ligne.get("departure_date")
                or ligne.get("order_date")
                or ""
            ).strip()
            try:
                date_vente = date.fromisoformat(texte_date[:10])
            except ValueError:
                continue
            if date_vente > date_coupure:
                raise RuntimeError(
                    f"Complement cadencier refuse: date hors coupure {date_vente}"
                )

            produits = produits_par_client[code_client]
            produit = produits.setdefault(
                code_article,
                {
                    "code_article": code_article,
                    "client_libelle": str(
                        ligne.get("client_label") or ""
                    ).strip(),
                    "libelle_article": libelle,
                    "libelle_normalise": normaliser(libelle),
                    "prix": None,
                    "nb_ventes_article_total": 0,
                    "nb_ventes_article_recentes": 0,
                    "derniere_vente_article_iso": "",
                    "derniere_vente_article_ordinal": -1,
                    "quantite_habituelle_commande": 0.0,
                    "quantites_habituelles_commande": [],
                    "ratio_net_par_unite": 0.0,
                    "unite_vente": str(ligne.get("unit") or "").strip().upper(),
                    "source_article": "historique_client_pretest",
                },
            )
            produit["client_libelle"] = (
                str(ligne.get("client_label") or "").strip()
                or produit.get("client_libelle", "")
            )
            produit["libelle_article"] = libelle
            produit["libelle_normalise"] = normaliser(libelle)
            produit["nb_ventes_article_total"] = (
                int(produit.get("nb_ventes_article_total", 0)) + 1
            )
            if date_vente >= seuil_recent:
                produit["nb_ventes_article_recentes"] = (
                    int(produit.get("nb_ventes_article_recentes", 0)) + 1
                )
            ordinal = date_vente.toordinal()
            if ordinal > int(produit.get("derniere_vente_article_ordinal", -1)):
                produit["derniere_vente_article_ordinal"] = ordinal
                produit["derniere_vente_article_iso"] = date_vente.isoformat()
            if not produit.get("unite_vente"):
                produit["unite_vente"] = str(
                    ligne.get("unit") or ""
                ).strip().upper()

            texte_quantite = str(ligne.get("quantity") or "").replace(",", ".")
            try:
                quantite = round(float(texte_quantite), 3)
            except ValueError:
                quantite = 0.0
            if quantite > 0:
                cle = (code_client, code_article)
                historiques_quantites[cle][quantite] += 1
                recence_quantites[cle][quantite] = max(
                    recence_quantites[cle].get(quantite, -1),
                    ordinal,
                )

    for (code_client, code_article), historique in historiques_quantites.items():
        produit = produits_par_client[code_client][code_article]
        quantites_tries = sorted(
            historique.items(),
            key=lambda item: (
                item[1],
                recence_quantites[(code_client, code_article)].get(item[0], -1),
                item[0],
            ),
            reverse=True,
        )
        produit["quantite_habituelle_commande"] = float(quantites_tries[0][0])
        produit["quantites_habituelles_commande"] = [
            float(quantite) for quantite, _ in quantites_tries[:5]
        ]


ALIASES_COLONNES_CADENCIER: dict[str, list[str]] = {
    "code_client": [
        "client code",
        "clients code",
        "client livre code",
        "clients livres code",
        "code client",
    ],
    "libelle_client": [
        "client lib",
        "clients lib",
        "client livre lib",
        "clients livres lib",
        "nom client",
    ],
    "code_article": ["article code", "code article"],
    "libelle_article": ["article lib", "libelle article"],
    "prix_net": [
        "mtt net pied livre",
        "mtt net pied",
        "px net pied",
        "prix net",
        "montant net",
    ],
    "quantite_livree": [
        "poids net livre",
        "poids net",
        "pieces livrees",
        "pieces liv",
        "pds net liv",
    ],
    "poids_net": ["poids net livre", "poids net", "pds net liv"],
    "date_depart": ["depart", "date depart", "date"],
    "numero_commande": ["n cde", "no cde", "numero commande"],
    "montant_net": [
        "mtt net pied livre",
        "mtt net pied",
        "mtt net ligne",
        "montant net ligne",
        "mtt net",
    ],
}


def charger_cadencier() -> dict[str, list[dict[str, Any]]]:
    chemin = trouver_fichier_xlsx(DOSSIER_CADENCIER)

    en_tetes, lignes = lire_lignes_xlsx(chemin)

    col_client = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["code_client"],
        champ_metier="code_client",
        chemin=chemin,
    )
    col_client_lib = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["libelle_client"],
        obligatoire=False,
        champ_metier="libelle_client",
        chemin=chemin,
    )

    col_article = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["code_article"],
        champ_metier="code_article",
        chemin=chemin,
    )

    col_libelle = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["libelle_article"],
        champ_metier="libelle_article",
        chemin=chemin,
    )
    col_prix = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["prix_net"],
        obligatoire=False,
        champ_metier="prix_net",
        chemin=chemin,
    )
    col_pieces = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["quantite_livree"],
        obligatoire=False,
        champ_metier="quantite_livree",
        chemin=chemin,
    )
    col_pds_net = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["poids_net"],
        obligatoire=False,
        champ_metier="poids_net",
        chemin=chemin,
    )
    col_depart = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["date_depart"],
        obligatoire=False,
        champ_metier="date_depart",
        chemin=chemin,
    )

    dates_valides = [
        valeur
        if isinstance(valeur, datetime)
        else datetime.combine(
            valeur,
            datetime.min.time(),
        )
        for valeur in (
            ligne.get(col_depart)
            for ligne in lignes
        )
        if col_depart
        and isinstance(valeur, (datetime, date))
    ]
    reference_recente = (
        max(dates_valides)
        if dates_valides
        else None
    )
    seuil_recent = (
        reference_recente - timedelta(days=30)
        if reference_recente is not None
        else None
    )

    produits_par_client: dict[
        str,
        dict[str, dict[str, Any]],
    ] = defaultdict(dict)

    for ligne in lignes:
        code_client = str(
            ligne.get(col_client, "")
            or ""
        ).strip()

        code_article = str(
            ligne.get(col_article, "")
            or ""
        ).strip()
        client_libelle = (
            str(ligne.get(col_client_lib, "") or "").strip()
            if col_client_lib
            else ""
        )

        libelle = str(
            ligne.get(col_libelle, "")
            or ""
        ).strip()
        prix = (
            ligne.get(col_prix)
            if col_prix
            else None
        )

        if not code_client or not code_article or not libelle:
            continue
        code_article = _canonicaliser_code_article_reference(
            code_article,
            libelle,
        )

        produit = produits_par_client[
            code_client
        ].setdefault(
            code_article,
            {
                "code_article": code_article,
                "client_libelle": client_libelle,
                "libelle_article": libelle,
                "libelle_normalise": normaliser(libelle),
                "prix": None,
                "nb_ventes_article_total": 0,
                "nb_ventes_article_recentes": 0,
                "derniere_vente_article_iso": "",
                "derniere_vente_article_ordinal": -1,
                "quantite_habituelle_commande": 0.0,
                "quantites_habituelles_commande": [],
                # Somme physique livree (poids/volume fourni par le
                # cadencier). Elle est distincte du nombre de lignes : trois
                # petites bouteilles ne doivent pas peser davantage qu'une
                # habitude de bidons de 10 L lors du departage.
                "volume_historique_total": 0.0,
                "ratio_net_par_unite": 0.0,
                "_historique_quantites": Counter(),
                "_recence_quantites": {},
                "_somme_quantites_positives": 0.0,
                "_somme_pds_net_positif": 0.0,
            },
        )

        produit["client_libelle"] = client_libelle or produit[
            "client_libelle"
        ]
        produit["libelle_article"] = libelle or produit[
            "libelle_article"
        ]
        produit["libelle_normalise"] = normaliser(
            produit["libelle_article"]
        )
        if isinstance(prix, (int, float)):
            produit["prix"] = float(prix)

        produit["nb_ventes_article_total"] += 1

        qte_pieces = (
            ligne.get(col_pieces)
            if col_pieces
            else None
        )
        pds_net = (
            ligne.get(col_pds_net)
            if col_pds_net
            else None
        )

        if col_depart:
            date_depart = ligne.get(col_depart)
            if isinstance(date_depart, datetime):
                date_depart_dt = date_depart
                date_depart_simple = date_depart.date()
            elif isinstance(date_depart, date):
                date_depart_simple = date_depart
                date_depart_dt = datetime.combine(
                    date_depart,
                    datetime.min.time(),
                )
            else:
                date_depart_simple = None
                date_depart_dt = None

            if date_depart_simple is not None:
                ordinal = date_depart_simple.toordinal()
                if (
                    ordinal
                    > produit["derniere_vente_article_ordinal"]
                ):
                    produit["derniere_vente_article_ordinal"] = (
                        ordinal
                    )
                    produit["derniere_vente_article_iso"] = (
                        date_depart_simple.isoformat()
                    )
            if (
                seuil_recent is not None
                and date_depart_dt is not None
                and date_depart_dt >= seuil_recent
            ):
                produit["nb_ventes_article_recentes"] += 1
            recence_ordinal = (
                int(date_depart_simple.toordinal())
                if date_depart_simple is not None
                else -1
            )
        else:
            produit["nb_ventes_article_recentes"] += 1
            recence_ordinal = -1

        if (
            isinstance(qte_pieces, (int, float))
            and float(qte_pieces) > 0
        ):
            qte_positive = round(
                float(qte_pieces), 3
            )
            produit["_historique_quantites"][
                qte_positive
            ] += 1
            produit["_recence_quantites"][
                qte_positive
            ] = max(
                int(
                    produit["_recence_quantites"].get(
                        qte_positive, -1
                    )
                ),
                recence_ordinal,
            )
            produit["_somme_quantites_positives"] += (
                qte_positive
            )

            if (
                isinstance(pds_net, (int, float))
                and float(pds_net) > 0
            ):
                produit["_somme_pds_net_positif"] += (
                    float(pds_net)
                )

        # Les exports BASCO recents exposent parfois uniquement le poids net
        # livre : cette valeur est alors a la fois la quantite observee et la
        # meilleure mesure de volume historique disponible.
        volume_livre = (
            float(pds_net)
            if isinstance(pds_net, (int, float)) and float(pds_net) > 0
            else (
                float(qte_pieces)
                if isinstance(qte_pieces, (int, float))
                and float(qte_pieces) > 0
                else 0.0
            )
        )
        produit["volume_historique_total"] += volume_livre

    for produits_client in produits_par_client.values():
        for produit in produits_client.values():
            historique_quantites = produit.pop(
                "_historique_quantites"
            )
            recence_quantites = produit.pop(
                "_recence_quantites"
            )
            somme_quantites = float(
                produit.pop(
                    "_somme_quantites_positives"
                )
            )
            somme_pds_net = float(
                produit.pop("_somme_pds_net_positif")
            )

            if historique_quantites:
                quantites_tries = sorted(
                    historique_quantites.items(),
                    key=lambda item: (
                        item[1],
                        recence_quantites.get(
                            item[0], -1
                        ),
                        item[0],
                    ),
                    reverse=True,
                )
                produit[
                    "quantite_habituelle_commande"
                ] = float(quantites_tries[0][0])
                produit[
                    "quantites_habituelles_commande"
                ] = [
                    float(quantite)
                    for quantite, _ in quantites_tries[:5]
                ]

            if somme_quantites > 0 and somme_pds_net > 0:
                produit["ratio_net_par_unite"] = round(
                    somme_pds_net / somme_quantites,
                    4,
                )

    # Le complement est derive de commandes ERP reelles. Il reste disponible
    # pour les analyses hors ligne, mais ne fait jamais partie des entrees du
    # moteur de prediction en mode d'evaluation sans fuite.

    return {
        code_client: list(produits.values())
        for code_client, produits in produits_par_client.items()
    }


def charger_stats_ventes_clients() -> dict[str, dict[str, Any]]:
    chemin = trouver_fichier_xlsx(DOSSIER_CADENCIER)
    en_tetes, lignes = lire_lignes_xlsx(chemin)

    col_client = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["code_client"],
        champ_metier="code_client",
        chemin=chemin,
    )
    col_depart = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["date_depart"],
        obligatoire=False,
        champ_metier="date_depart",
        chemin=chemin,
    )
    col_commande = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["numero_commande"],
        obligatoire=False,
        champ_metier="numero_commande",
        chemin=chemin,
    )
    col_montant = choisir_colonne(
        en_tetes,
        ALIASES_COLONNES_CADENCIER["montant_net"],
        obligatoire=False,
        champ_metier="montant_net",
        chemin=chemin,
    )

    dates_valides = [
        valeur
        if isinstance(valeur, datetime)
        else datetime.combine(
            valeur,
            datetime.min.time(),
        )
        for valeur in (
            ligne.get(col_depart)
            for ligne in lignes
        )
        if col_depart
        and isinstance(valeur, (datetime, date))
    ]
    reference_recente = (
        max(dates_valides)
        if dates_valides
        else None
    )
    seuil_recent = (
        reference_recente - timedelta(days=30)
        if reference_recente is not None
        else None
    )

    stats_par_client: dict[str, dict[str, Any]] = {}

    for ligne in lignes:
        code_client = str(
            ligne.get(col_client, "") or ""
        ).strip()
        date_depart = ligne.get(col_depart) if col_depart else None

        if not code_client:
            continue

        stats = stats_par_client.setdefault(
            code_client,
            {
                "derniere_vente_iso": "",
                "derniere_vente_ordinal": -1,
                "nb_lignes_ventes": 0,
                "nb_lignes_ventes_recentes": 0,
                "nb_commandes_total": 0,
                "nb_commandes_recentes": 0,
                "montant_recent": 0.0,
                "_commandes_total": set(),
                "_commandes_recentes": set(),
            },
        )

        stats["nb_lignes_ventes"] += 1
        if date_depart:
            if isinstance(date_depart, datetime):
                date_depart_dt = date_depart
                date_depart_simple = date_depart.date()
            elif isinstance(date_depart, date):
                date_depart_simple = date_depart
                date_depart_dt = datetime.combine(
                    date_depart,
                    datetime.min.time(),
                )
            else:
                date_depart_simple = None
                date_depart_dt = None

            if date_depart_simple is not None:
                ordinal = date_depart_simple.toordinal()
                if (
                    ordinal
                    > stats["derniere_vente_ordinal"]
                ):
                    stats["derniere_vente_ordinal"] = (
                        ordinal
                    )
                    stats["derniere_vente_iso"] = (
                        date_depart_simple.isoformat()
                    )
            est_recent = (
                seuil_recent is not None
                and date_depart_dt is not None
                and date_depart_dt >= seuil_recent
            )
        else:
            est_recent = True

        if est_recent:
            stats["nb_lignes_ventes_recentes"] += 1

        numero_commande = (
            str(ligne.get(col_commande, "") or "").strip()
            if col_commande
            else ""
        )
        if numero_commande:
            stats["_commandes_total"].add(numero_commande)
            if est_recent:
                stats["_commandes_recentes"].add(
                    numero_commande
                )
        montant = (
            ligne.get(col_montant)
            if col_montant
            else None
        )
        if isinstance(montant, (int, float)):
            stats["montant_recent"] += float(montant)

    for stats in stats_par_client.values():
        stats["nb_commandes_total"] = len(
            stats.pop("_commandes_total")
        )
        stats["nb_commandes_recentes"] = len(
            stats.pop("_commandes_recentes")
        )
        stats["montant_recent"] = round(
            stats["montant_recent"], 2
        )

    return stats_par_client


def enrichir_clients_depuis_cadencier(
    clients: list[dict[str, Any]],
    cadencier: dict[str, list[dict[str, Any]]],
) -> int:
    clients_par_code = {
        client["code_client"]: client
        for client in clients
    }
    ajoutes = 0

    for code_client, produits in cadencier.items():
        if code_client in clients_par_code:
            continue

        nom_client = ""

        if produits:
            nom_client = str(
                produits[0].get("client_libelle", "")
                or ""
            ).strip()

        if not nom_client:
            nom_client = code_client

        ville = ""
        tokens_nom = normaliser(nom_client).split()

        if tokens_nom:
            token_final = tokens_nom[-1]

            if (
                token_final.isalpha()
                and len(token_final) >= 4
            ):
                ville = token_final.upper()

        aliases = set(
            creer_alias_client(nom_client)
        )
        aliases.add(normaliser(code_client))

        nouveau_client = {
            "code_client": code_client,
            "nom_client": nom_client,
            "ville": ville,
            "adresse_1": "",
            "adresse_2": "",
            "code_postal": "",
            "aliases": sorted(
                alias
                for alias in aliases
                if alias
            ),
            "source_client": "cadencier",
        }

        clients.append(nouveau_client)
        clients_par_code[code_client] = nouveau_client
        ajoutes += 1

    return ajoutes


def enrichir_clients_avec_stats_ventes(
    clients: list[dict[str, Any]],
    stats_ventes: dict[str, dict[str, Any]],
) -> None:
    for client in clients:
        stats = stats_ventes.get(
            str(client.get("code_client", "")).strip(),
            {},
        )
        client["derniere_vente_iso"] = stats.get(
            "derniere_vente_iso",
            "",
        )
        client["derniere_vente_ordinal"] = int(
            stats.get("derniere_vente_ordinal", -1)
        )
        client["nb_lignes_ventes"] = int(
            stats.get("nb_lignes_ventes", 0)
        )
        client["nb_lignes_ventes_recentes"] = int(
            stats.get("nb_lignes_ventes_recentes", 0)
        )
        client["nb_commandes_total"] = int(
            stats.get("nb_commandes_total", 0)
        )
        client["nb_commandes_recentes"] = int(
            stats.get("nb_commandes_recentes", 0)
        )
        client["montant_recent"] = float(
            stats.get("montant_recent", 0.0)
        )


# -------------------------------------------------------------------
# Dates
# -------------------------------------------------------------------

def extraire_date_livraison(
    transcription: str,
    date_reference: date | None = None,
) -> dict[str, str] | None:
    """
    Détection volontairement stricte.

    Elle évite les faux positifs comme :
    'je' interprété comme une date.
    """

    reference = date_reference or date.today()

    texte = normaliser(transcription)
    candidats: list[
        tuple[int, int, dict[str, str]]
    ] = []

    for motif in re.finditer(
        r"\bapres demain\b",
        texte,
    ):
        valeur = reference + timedelta(days=2)
        candidats.append(
            (
                motif.start(),
                motif.end(),
                {
                    "expression": "après-demain",
                    "date_iso": valeur.isoformat(),
                },
            )
        )

    for motif in re.finditer(
        r"\bdemain\b",
        texte,
    ):
        valeur = reference + timedelta(days=1)
        candidats.append(
            (
                motif.start(),
                motif.end(),
                {
                    "expression": "demain",
                    "date_iso": valeur.isoformat(),
                },
            )
        )

    for motif_date in re.finditer(
        r"\b(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?\b",
        texte,
    ):
        jour = int(motif_date.group(1))
        mois = int(motif_date.group(2))

        annee_brute = motif_date.group(3)

        if annee_brute is None:
            annee = reference.year
        elif len(annee_brute) == 2:
            annee = 2000 + int(annee_brute)
        else:
            annee = int(annee_brute)

        try:
            valeur = date(
                year=annee,
                month=mois,
                day=jour,
            )
        except ValueError:
            continue

        candidats.append(
            (
                motif_date.start(),
                motif_date.end(),
                {
                    "expression": motif_date.group(0),
                    "date_iso": valeur.isoformat(),
                },
            )
        )

    for mot, numero_jour in JOURS_SEMAINE.items():
        for motif in re.finditer(
            rf"\b{re.escape(mot)}\b",
            texte,
        ):
            decalage = (
                numero_jour
                - reference.weekday()
            ) % 7

            if decalage == 0:
                decalage = 7

            valeur = reference + timedelta(
                days=decalage
            )

            candidats.append(
                (
                    motif.start(),
                    motif.end(),
                    {
                        "expression": mot,
                        "date_iso": valeur.isoformat(),
                    },
                )
            )

    if not candidats:
        return None

    _, _, premier = min(
        candidats,
        key=lambda candidat: (
            candidat[0],
            -(candidat[1] - candidat[0]),
        ),
    )

    return premier


def _message_est_rappel_sans_commande(
    texte: str,
) -> bool:
    if not any(
        expression in texte
        for expression in MOTIFS_RAPPEL_CLIENT
    ):
        return False

    marqueurs_commande = (
        "je voudrais",
        "il me faudrait",
        "il faudrait",
        "commande",
        "ajouter",
        "ajoute",
    )
    if any(
        marqueur in texte
        for marqueur in marqueurs_commande
    ):
        return False

    unites = (
        "carton",
        "cartons",
        "kilo",
        "kilos",
        "kg",
        "litre",
        "litres",
        "piece",
        "pieces",
        "pack",
        "packs",
        "bidon",
        "bidons",
        "boite",
        "boites",
    )
    mots_nombres = tuple(NOMBRES.keys())
    motif_quantite = (
        r"\b(?:\d+(?:[\.,]\d+)?|"
        + "|".join(
            re.escape(mot) for mot in mots_nombres
        )
        + r")\s+(?:"
        + "|".join(re.escape(unite) for unite in unites)
        + r")\b"
    )

    return re.search(motif_quantite, texte) is None


def detecter_type_action_commande(
    transcription: str,
) -> dict[str, str]:
    texte = normaliser(transcription)

    for motif_modification in MOTIFS_MODIFICATION_A_RAPPELER:
        motif = re.search(motif_modification, texte)
        if motif:
            return {
                "type_action": "modification",
                "expression": motif.group(0),
            }

    for expression in MOTIFS_SUPPRESSION_COMMANDE:
        motif = re.search(
            rf"\b{re.escape(expression)}\b",
            texte,
        )
        if motif:
            return {
                "type_action": "suppression",
                "expression": expression,
            }

    if _message_est_rappel_sans_commande(texte):
        return {
            "type_action": "rappel",
            "expression": "rappel_client",
        }

    for expression in MOTIFS_NOUVELLE_COMMANDE:
        motif = re.search(
            rf"\b{re.escape(expression)}\b",
            texte,
        )
        if motif:
            return {
                "type_action": "creation",
                "expression": expression,
            }

    return {
        "type_action": "creation",
        "expression": "",
    }


def extraire_mentions_suppression(
    transcription: str,
) -> list[dict[str, Any]]:
    texte = normaliser(transcription)
    segments: list[str] = []

    for verbe in MOTIFS_SUPPRESSION_COMMANDE:
        for motif in re.finditer(
            rf"\b{re.escape(verbe)}\b\s+(?P<segment>.+)",
            texte,
        ):
            segment = motif.group("segment").strip()
            segment = re.split(
                r"\b(?:pour|merci|bonne journee|au revoir|"
                r"ce sera|ca sera|ce serait|ca serait)\b",
                segment,
                maxsplit=1,
            )[0].strip()

            if segment:
                segments.append(segment)

    mentions: list[dict[str, Any]] = []

    for segment in segments:
        morceaux = re.split(
            r"\s*(?:,|;|\bet\b)\s*",
            segment,
        )

        for morceau in morceaux:
            produit = re.sub(
                r"^(?:le|la|les|du|de la|de l|des|d)\s+",
                "",
                morceau.strip(),
            )
            produit = produit.strip()

            if len(produit) < 3:
                continue

            mentions.append(
                {
                    "texte_source": produit,
                    "texte_normalise": produit,
                    "produit_normalise": produit,
                    "texte_produit": produit,
                    "quantite_principale": 1.0,
                    "quantite": 1.0,
                    "unite_principale": None,
                    "unite_detectee": None,
                    "precisions_quantite": [],
                    "ambigu": False,
                    "raisons_ambiguite": [],
                }
            )

    return mentions


def resoudre_date_livraison(
    transcription: str,
    date_reference: date | None = None,
    heure_reference: int | None = None,
    jour_reference: date | None = None,
) -> dict[str, Any]:
    date_expression_reference = date_reference
    if (
        date_expression_reference is not None
        and heure_reference is not None
        and 0 <= heure_reference < 6
    ):
        date_expression_reference -= timedelta(days=1)

    date_detectee = extraire_date_livraison(
        transcription=transcription,
        date_reference=date_expression_reference,
    )

    if date_detectee:
        return {
            **date_detectee,
            "date_par_defaut": False,
        }

    if date_reference is not None and heure_reference is not None:
        if 0 <= heure_reference < 6:
            # Appels en soirée (>=22h) ou tôt le matin (<6h) : livraison le jour même.
            # Les clients passent commande tard le soir pour le lendemain matin,
            # qu'ES enregistre avec order_date = date du jour de l'appel.
            reference = date_reference
            expression = "defaut_nuit_date_du_jour"
        else:
            reference = date_reference + timedelta(days=1)
            expression = (
                "defaut_soiree_date_demain"
                if heure_reference >= 22
                else "defaut_journee_date_demain"
            )
    elif date_reference is not None:
        reference = date_reference
        expression = "defaut_date_reference"
    else:
        heure = (
            heure_reference
            if heure_reference is not None
            else datetime.now().hour
        )
        jour = jour_reference or date.today()

        if 0 <= heure < 6:
            reference = jour
            expression = "defaut_nuit_date_du_jour"
        else:
            reference = jour + timedelta(days=1)
            expression = (
                "defaut_soiree_date_demain"
                if heure >= 22
                else "defaut_journee_date_demain"
            )

    return {
        "expression": expression,
        "date_iso": reference.isoformat(),
        "date_par_defaut": True,
    }


# -------------------------------------------------------------------
# Extraction des mentions produit
# -------------------------------------------------------------------

def remplacer_nombres_ecrits(texte: str) -> str:
    # Conservé pour compatibilité avec les tests
    # et l’ancien flux d’appel.
    return texte


def nettoyer_clause(clause: str) -> str:
    return clause.strip()


def decouper_clauses_produits(
    transcription: str,
) -> list[str]:
    return decouper_clauses_produits_v2(
        transcription
    )


def extraire_mentions_produits(
    transcription: str,
) -> list[dict[str, Any]]:
    return extraire_mentions_produits_v2(
        transcription
    )


# -------------------------------------------------------------------
# Recherche des produits dans le cadencier
# -------------------------------------------------------------------

def chercher_produits(
    mentions: list[dict[str, Any]],
    produits_client: list[dict[str, Any]],
    catalogue_global: list[dict[str, Any]],
    synonymes_produits: dict[str, list[str]],
    catalogue_reappro: list[dict[str, Any]] = None,
    limite: int = 50,
) -> list[dict[str, Any]]:
    produits = chercher_produits_v2(
        mentions=mentions,
        produits_client=produits_client,
        catalogue_global=catalogue_global,
        catalogue_reappro=catalogue_reappro,
        synonymes=synonymes_produits,
        limite=limite,
    )

    # Compatibilité avec les anciens appels : clef `score`.
    for produit in produits:
        for candidat in produit.get("candidats", []):
            candidat["score"] = candidat["score_global"]

    # --- Arbitrage LLM pour les mentions phonétiques ambiguës ---
    # Si un produit n'a pas de candidat avec un score suffisant (< 55),
    # on demande au LLM local de trancher parmi les candidats disponibles.
    try:
        from src.llm_arbitrage import arbitrer_produit_phonetique, ollama_disponible
        _llm_ok = ollama_disponible()
    except Exception:
        _llm_ok = False

    if _llm_ok:
        for produit in produits:
            candidats = produit.get("candidats", [])
            if not candidats:
                continue
            meilleur_score = max(
                (c.get("score_global", 0) for c in candidats), default=0
            )
            # On n'intervient que si le moteur classique est incertain
            if meilleur_score < 55.0:
                mention_texte = produit.get("texte_source", produit.get("texte_produit", ""))
                if not mention_texte:
                    continue
                candidats_pour_llm = [
                    {
                        "code_article": c.get("code_article", ""),
                        "libelle": c.get("libelle_normalise", c.get("libelle", "")),
                    }
                    for c in candidats[:15]
                ]
                choix = arbitrer_produit_phonetique(
                    mention_texte=mention_texte,
                    candidats=candidats_pour_llm,
                )
                if choix:
                    # On retrouve le candidat complet correspondant au choix du LLM
                    code_choisi = choix.get("code_article", "")
                    for c in candidats:
                        if c.get("code_article", "") == code_choisi:
                            c["score_global"] = max(c.get("score_global", 0), 60.0)
                            c["score"] = c["score_global"]
                            c["llm_arbitrage"] = True
                            break

    return produits



# -------------------------------------------------------------------
# Lecture et écriture JSON / TXT
# -------------------------------------------------------------------

def lire_transcription(
    chemin_json: Path,
) -> str:
    donnees = json.loads(
        chemin_json.read_text(
            encoding="utf-8"
        )
    )

    return str(
        donnees.get(
            "texte",
            "",
        )
    ).strip()


def creer_resume_txt(
    resultat: dict[str, Any],
) -> str:
    lignes = [
        f"AUDIO : {resultat['fichier_audio']}",
        "",
        "TRANSCRIPTION :",
        resultat["transcription"],
        "",
        "CLIENTS CANDIDATS :",
    ]

    for candidat in resultat["clients_candidats"]:
        score_global = candidat.get("score_global")
        score_nom = candidat.get("score_nom")
        score_code = candidat.get("score_code")
        score_adresse = candidat.get("score_adresse")
        score_ville = candidat.get("score_ville")
        score_cadencier = candidat.get(
            "score_cadencier"
        )
        score_ancien = candidat.get("score")

        details_scores: list[str] = []

        if score_global is not None:
            details_scores.append(
                f"global={score_global}"
            )

        if score_nom is not None:
            details_scores.append(
                f"nom={score_nom}"
            )

        if score_code is not None:
            details_scores.append(
                f"code={score_code}"
            )

        if score_adresse is not None:
            details_scores.append(
                f"adresse={score_adresse}"
            )

        if score_ville is not None:
            details_scores.append(
                f"ville={score_ville}"
            )

        if score_cadencier is not None:
            details_scores.append(
                f"cadencier={score_cadencier}"
            )

        if (
            not details_scores
            and score_ancien is not None
        ):
            details_scores.append(
                f"score={score_ancien}"
            )

        lignes.append(
            "- "
            f"{candidat['code_client']} "
            "| "
            f"{candidat['nom_client']} "
            "| "
            + " / ".join(details_scores)
        )

    if not resultat["clients_candidats"]:
        lignes.append("- Aucun client candidat")

    lignes.extend(
        [
            "",
            "ACTION COMMANDE :",
            str(
                resultat.get(
                    "type_action_commande",
                    "creation",
                )
            ),
            "",
            (
                "CLIENT RETENU AUTOMATIQUEMENT : "
                f"{resultat['client_retenu'] or 'NON'}"
            ),
            (
                "DÉCISION AUTOMATIQUE : "
                f"{'OUI' if resultat.get('decision_automatique_client') else 'NON'}"
            ),
            (
                "RAISONS DÉCISION CLIENT : "
                f"{resultat.get('raisons_decision_client', [])}"
            ),
            "",
            "DATE DE LIVRAISON :",
            str(
                resultat["date_livraison"]
                or "Non détectée"
            ),
            "",
            "PRODUITS :",
        ]
    )

    if not resultat["produits"]:
        lignes.append("- Aucun produit détecté")

    for produit in resultat["produits"]:
        quantite_affichee = (
            produit.get("quantite_resolue")
            if produit.get("quantite_resolue")
            is not None
            else produit.get("quantite")
        )
        unite_affichee = (
            produit.get("unite_resolue")
            or produit.get("unite_detectee")
            or "?"
        )
        lignes.append(
            "- Mention : "
            f"{quantite_affichee} "
            f"{unite_affichee} "
            f"de {produit['texte_produit']}"
        )

        if produit["precisions_quantite"]:
            lignes.append(
                "  Précisions de quantité : "
                f"{produit['precisions_quantite']}"
            )

        for candidat in produit["candidats"]:
            lignes.append(
                "    * "
                f"{candidat['code_article']} "
                "| "
                f"{candidat['libelle_article']} "
                f"| score={candidat['score']}"
            )

        lignes.append(
            "  Sélection automatique fiable : "
            + (
                "OUI"
                if produit["produit_fiable"]
                else "NON"
            )
        )

    return "\n".join(lignes)


# -------------------------------------------------------------------
# Commandes / CSV
# -------------------------------------------------------------------

# Les fonctions produit publiques sont importees une seule fois en tete de
# module. Les wrappers de compatibilite ci-dessus doivent rester actifs : une
# seconde importation ici les ecrasait et referencait en plus des symboles qui
# n'existent pas dans ``src.produits``, rendant tout le moteur inimportable.

@lru_cache(maxsize=1)
def charger_catalogue_articles_reference() -> list[dict[str, Any]]:
    """Charge le referentiel produit local, independant des commandes evaluees."""
    if not CHEMIN_CATALOGUE_ARTICLES.exists():
        return []
    try:
        payload = json.loads(CHEMIN_CATALOGUE_ARTICLES.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return [item for item in payload if isinstance(item, dict)]

@lru_cache(maxsize=1)
def charger_reappro_basco() -> list[dict[str, Any]]:
    """Charge le referentiel complementaire Reappro BASCO."""
    CHEMIN_REAPPRO = DOSSIER_CONFIG / "reappro_basco.csv"
    if not CHEMIN_REAPPRO.exists():
        return []
        
    references = {}
    if CHEMIN_REFERENCES_CONTROLE.exists():
        try:
            ref_payload = json.loads(CHEMIN_REFERENCES_CONTROLE.read_text(encoding="utf-8"))
            if isinstance(ref_payload, dict):
                references = ref_payload.get("references", {})
        except (OSError, json.JSONDecodeError):
            pass

    payload = []
    try:
        import csv
        with CHEMIN_REAPPRO.open("r", encoding="latin-1", newline="") as fichier:
            reader = csv.reader(fichier, delimiter=";")
            headers = next(reader, [])
            if "Produit code" in headers:
                idx_code = headers.index("Produit code")
                idx_lib = headers.index("Produit lib") if "Produit lib" in headers else -1
                
                codes_vus = set()
                for row in reader:
                    if not row or len(row) <= idx_code:
                        continue
                    code = row[idx_code].strip()
                    if not code or code in codes_vus:
                        continue
                    codes_vus.add(code)
                    
                    libelle = ""
                    if code in references and isinstance(references[code], dict) and references[code].get("label"):
                        libelle = str(references[code].get("label")).strip()
                    elif idx_lib >= 0 and len(row) > idx_lib:
                        libelle = row[idx_lib].strip()
                        
                    payload.append({
                        "code_article": code,
                        "libelle_article": libelle,
                        "libelle_normalise": normaliser(libelle),
                        "source_article": "referentiel_articles",
                        "prix": 0.01,
                    })
    except OSError:
        pass
        
    return payload


@lru_cache(maxsize=1)
def charger_unites_articles() -> dict[str, str]:
    """Charge les unités officielles, puis les rares ajustements locaux."""
    unites: dict[str, str] = {}
    if CHEMIN_REFERENCES_CONTROLE.exists():
        try:
            payload = json.loads(
                CHEMIN_REFERENCES_CONTROLE.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            payload = {}
        for code, reference in (
            payload.get("references", {}).items()
            if isinstance(payload, dict)
            else []
        ):
            if not isinstance(reference, dict):
                continue
            unite = str(reference.get("order_unit") or "").strip().upper()
            if code and unite:
                unites[str(code)] = unite

    if CHEMIN_UNITES_ARTICLES.exists():
        with CHEMIN_UNITES_ARTICLES.open(
            "r", encoding="utf-8-sig", newline=""
        ) as fichier:
            for ligne in csv.DictReader(fichier, delimiter=";"):
                code = str(ligne.get("code_article") or "").strip()
                unite = str(ligne.get("unite") or "").strip().upper()
                if code and unite:
                    unites[code] = unite
    return unites


@lru_cache(maxsize=1)
def charger_commandes_courtes_clients() -> list[dict[str, Any]]:
    """Charge les raccourcis client explicitement valides par le metier."""
    if not CHEMIN_COMMANDES_COURTES_CLIENTS.exists():
        return []
    contenu = json.loads(
        CHEMIN_COMMANDES_COURTES_CLIENTS.read_text(
            encoding="utf-8"
        )
    )
    if not isinstance(contenu, dict):
        return []
    commandes = contenu.get("commandes_courtes", [])
    return [
        commande
        for commande in commandes
        if isinstance(commande, dict)
    ]


def appliquer_commande_courte_client(
    *,
    transcription: str,
    client_code: str | None,
    type_action: str,
    mentions: list[dict[str, Any]],
    produits: list[dict[str, Any]],
    produits_client: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Transforme un raccourci client valide en une ligne tracee.

    Ce mecanisme est volontairement declaratif et tres borne : il ne se
    declenche ni sur une modification, ni si l'appel contient deja une
    quantite produit. Il n'apprend jamais seul une commande implicite.
    """
    if not client_code or type_action != "creation":
        return produits, None
    if any(
        mention.get("quantite_principale") is not None
        for mention in mentions
    ):
        return produits, None
    if any(
        produit.get("selection")
        and produit.get("quantite_resolue") is not None
        for produit in produits
    ):
        return produits, None

    texte_normalise = normaliser(transcription)
    for commande in charger_commandes_courtes_clients():
        if str(commande.get("client_code") or "") != client_code:
            continue
        triggers = [
            normaliser(trigger)
            for trigger in commande.get("triggers", [])
            if normaliser(trigger)
        ]
        occurrences_min = int(
            commande.get("occurrences_min", 1) or 1
        )
        if not triggers or sum(
            texte_normalise.count(trigger)
            for trigger in triggers
        ) < occurrences_min:
            continue

        code_article = str(commande.get("code_article") or "")
        article = next(
            (
                produit
                for produit in produits_client
                if str(produit.get("code_article") or "")
                == code_article
            ),
            None,
        )
        if article is None:
            # Une preference ne peut jamais commander une reference absente
            # du cadencier courant du client.
            continue

        quantite = float(commande.get("quantite") or 0.0)
        unite = str(commande.get("unite") or "").upper()
        if quantite <= 0 or not unite:
            continue

        libelle = str(article.get("libelle_article") or "")
        segment_id = (
            "commande-courte-"
            f"{client_code.lower()}-{code_article.lower()}"
        )
        produit_configure = {
            "segment_id": segment_id,
            "segment_index": 1,
            "texte_source": "commande courte client configuree",
            "produit_normalise": normaliser(libelle),
            "texte_produit": normaliser(libelle),
            "quantite": quantite,
            "quantite_principale": quantite,
            "unite_principale": unite,
            "unite_detectee": unite,
            "precisions_quantite": [],
            "conditionnement_multiple": None,
            "quantite_resolue": quantite,
            "unite_resolue": unite,
            "produit_fiable": True,
            "ambigu": False,
            "raisons_ambiguite": [],
            "candidats": [],
            "selection": {
                "code_article": code_article,
                "libelle_article": libelle,
                "score_texte": 100.0,
                "score_global": 100.0,
                "score_selection": 100.0,
                "source_recherche": "commande_courte_client_configuree",
                "prix": article.get("prix"),
            },
        }
        return [produit_configure], {
            "appliquee": True,
            "client_code": client_code,
            "code_article": code_article,
            "quantite": quantite,
            "unite": unite,
            "raison": "commande_courte_client_configuree",
        }

    return produits, None


def arbitrer_produits_ambigus_llama(
    produits: list[dict[str, Any]],
    client_nom: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Arbitre les selections faibles parmi des candidats actifs bornes.

    Les choix lexicalement forts restent exclusivement deterministes. Llama
    intervient seulement lorsqu'un prior cadencier peut avoir departage a tort
    des alternatives locales, ou lors d'une collision de codes deja connue.
    """
    try:
        from src.llm_arbitrage import (
            arbitrer_produit_phonetique,
            ollama_disponible,
        )
    except Exception:
        return produits, []
    if not ollama_disponible():
        return produits, []

    # Deux garde-fous bornent le cout et le perimetre de l'arbitrage : jamais
    # plus de cinq vraies mentions quantifiees par audio, et jamais le
    # catalogue global entier. Les candidats restent cadencier/Reapro.
    max_arbitrages = 5
    seuil_selection_faible = 65.0
    seuil_protection_deterministe = 60.0
    seuil_candidat_lexical = 35.0
    ecart_texte_max = 30.0
    occurrences_codes = Counter(
        str((produit.get("selection") or {}).get("code_article") or "")
        for produit in produits
        if produit.get("quantite_principale") is not None
        and str(
            (produit.get("selection") or {}).get("code_article") or ""
        )
    )
    audits: list[dict[str, Any]] = []
    nb_arbitrages = 0
    for produit in produits:
        if nb_arbitrages >= max_arbitrages:
            break
        selection = produit.get("selection") or {}
        # Une quantite explicite protege des formules de politesse, noms de
        # client et phrases de livraison. La segmentation reste le premier
        # filtre ; Llama n'est jamais utilise pour en creer une nouvelle.
        if produit.get("quantite_principale") is None:
            continue

        candidats_actifs = [
            candidat
            for candidat in produit.get("candidats", [])
            if candidat.get("source_recherche")
            in {"cadencier_client", "catalogue_reappro"}
            and candidat.get("semantiquement_compatible", True)
        ]
        if len(candidats_actifs) < 2:
            continue

        code_initial = str(selection.get("code_article") or "")
        score_selection = float(selection.get("score_texte") or 0.0)
        collision_faible = bool(
            code_initial
            and score_selection < 60.0
            and occurrences_codes.get(code_initial, 0) >= 2
        )
        meilleur_score_texte = max(
            float(candidat.get("score_texte") or 0.0)
            for candidat in candidats_actifs
        )
        candidats_lexicaux = [
            candidat
            for candidat in candidats_actifs
            if float(candidat.get("score_texte") or 0.0)
            >= max(
                seuil_candidat_lexical,
                meilleur_score_texte - ecart_texte_max,
            )
        ]
        alternative_lexicale = any(
            str(candidat.get("code_article") or "") != code_initial
            for candidat in candidats_lexicaux
        )
        selection_faible_avec_alternative = bool(
            score_selection < seuil_selection_faible
            and len(candidats_lexicaux) >= 2
            and alternative_lexicale
        )
        if not collision_faible and not selection_faible_avec_alternative:
            continue

        candidats = (
            candidats_actifs[:8]
            if collision_faible
            else candidats_lexicaux[:8]
        )
        if len(candidats) < 2:
            continue

        declencheur = (
            "collision_code_faible"
            if collision_faible
            else "selection_faible_alternative_lexicale"
        )

        choix = arbitrer_produit_phonetique(
            mention_texte=str(produit.get("texte_source") or ""),
            candidats=candidats,
            client_nom=client_nom,
        )
        nb_arbitrages += 1
        if not choix:
            audit = {
                "segment_id": produit.get("segment_id", ""),
                "applique": False,
                "raison": "llama_sans_choix",
                "declencheur": declencheur,
                "candidats_envoyes": len(candidats),
            }
            # Sur une selection isolee faible, une absence de choix Llama ne
            # doit pas laisser une substitution hors famille devenir une ligne
            # automatiquement commandable. Les collisions existantes gardent
            # leur comportement historique pour ne pas casser leur correctif.
            if selection_faible_avec_alternative and not collision_faible:
                produit["produit_fiable"] = False
                produit["produit_reconnu"] = False
                produit["ambigu"] = True
                if produit.get("statut_couverture") != "HORS_COMMANDE":
                    produit["statut_couverture"] = "NON_IDENTIFIE"
                produit.setdefault("raisons_ambiguite", []).append(
                    "llm_produit_sans_choix"
                )
                audit["raison"] = "llama_sans_choix_ligne_a_confirmer"
            audits.append(audit)
            continue

        code_choisi = str(choix.get("code_article") or "")
        if not code_choisi or code_choisi == code_initial:
            audits.append({
                "segment_id": produit.get("segment_id", ""),
                "applique": False,
                "raison": "llama_confirme_choix_initial",
                "declencheur": declencheur,
                "candidats_envoyes": len(candidats),
                "code_article": code_initial,
            })
            continue

        # Le catalogue de secours ne doit pas court-circuiter un article de
        # la meme famille deja plausible dans le cadencier client. Il reste
        # autorise lorsqu'il apporte une information explicite nettement plus
        # forte (variante semantique, conditionnement ou avantage lexical
        # important). Cette regle ne connait aucun code article particulier.
        choix_dans_cadencier = bool(
            choix.get("dans_cadencier_client")
        )
        cadencier_plausible = [
            candidat
            for candidat in candidats_actifs
            if candidat.get("dans_cadencier_client")
            and candidat.get("semantiquement_compatible", True)
            and (
                candidat.get("noyau_eligible_signaux_secondaires")
                or float(candidat.get("score_texte") or 0.0) >= 55.0
            )
        ]
        if not choix_dans_cadencier and cadencier_plausible:
            meilleur_cadencier = max(
                cadencier_plausible,
                key=lambda candidat: float(
                    candidat.get("score_texte") or 0.0
                ),
            )
            gain_texte = (
                float(choix.get("score_texte") or 0.0)
                - float(meilleur_cadencier.get("score_texte") or 0.0)
            )
            gain_attribut = (
                float(choix.get("score_attribut_semantique") or 0.0)
                - float(
                    meilleur_cadencier.get(
                        "score_attribut_semantique"
                    ) or 0.0
                )
            )
            gain_conditionnement = (
                float(
                    choix.get("score_conditionnement_physique_sur") or 0.0
                )
                - float(
                    meilleur_cadencier.get(
                        "score_conditionnement_physique_sur"
                    ) or 0.0
                )
            )
            secours_explicitement_justifie = bool(
                choix.get("noyau_eligible_signaux_secondaires")
                and (
                    gain_texte >= 20.0
                    or gain_attribut >= 15.0
                    or gain_conditionnement >= 10.0
                )
            )
            if not secours_explicitement_justifie:
                audits.append({
                    "segment_id": produit.get("segment_id", ""),
                    "applique": False,
                    "raison": "llama_secours_rejete_cadencier_plausible",
                    "declencheur": declencheur,
                    "code_initial": code_initial,
                    "code_article": code_choisi,
                    "code_cadencier_protege": meilleur_cadencier.get(
                        "code_article", ""
                    ),
                })
                continue

        # Llama sert a lever une ambiguite, pas a ecarter arbitrairement une
        # selection deterministe deja plus proche de la mention. Une
        # exception reste possible quand un signal semantique deterministe
        # etablit que le premier candidat est incompatible (ex. fruits rouges
        # versus fruits exotiques). Ce garde-fou est volontairement fonde sur
        # les scores et attributs des candidats, jamais sur un code article.
        score_initial = float(selection.get("score_texte") or 0.0)
        score_choisi = float(choix.get("score_texte") or 0.0)
        attribut_initial = float(
            selection.get("score_attribut_semantique") or 0.0
        )
        attribut_choisi = float(
            choix.get("score_attribut_semantique") or 0.0
        )
        remplacement_semantiquement_justifie = (
            attribut_choisi > attribut_initial
        )
        remplacement_moins_lexicalement_etaye = (
            score_choisi < score_initial
        )
        if (
            remplacement_moins_lexicalement_etaye
            and score_initial >= seuil_protection_deterministe
            and not remplacement_semantiquement_justifie
        ):
            audits.append({
                "segment_id": produit.get("segment_id", ""),
                "applique": False,
                "raison": "llama_remplacement_rejete_protection_deterministe",
                "declencheur": declencheur,
                "code_initial": code_initial,
                "code_article": code_choisi,
                "score_texte_initial": score_initial,
                "score_texte_choisi": score_choisi,
            })
            continue

        # Llama peut departager des candidats, mais il ne peut pas fabriquer
        # la preuve qu'une clause est un produit. Le choix doit encore passer
        # le meme product-gate deterministe que le moteur normal.
        noyau_llama_prouve, raisons_noyau_llama = (
            _preuve_positive_noyau_produit(
                str(produit.get("texte_source") or ""),
                choix,
                list(produit.get("variantes_recherche") or []),
                produit,
            )
        )
        tokens_source_llama = [
            token
            for token in normaliser(
                str(produit.get("texte_source") or "")
            ).split()
            if len(token) >= 5 and not token.isdigit()
        ]
        tokens_libelle_llama = [
            token
            for token in normaliser(
                str(choix.get("libelle_article") or "")
            ).split()
            if len(token) >= 5 and not token.isdigit()
        ]
        score_phonetique_llama = max(
            (
                float(fuzz.ratio(source, libelle))
                for source in tokens_source_llama
                for libelle in tokens_libelle_llama
            ),
            default=0.0,
        )
        # Ce secours autorise une deformation ASR substantielle que Llama sait
        # lever (nom etranger, suffixe parasite), sans accepter un choix sans
        # aucune proximite comme ``au moinsin`` -> ``cognac``.
        if (
            not noyau_llama_prouve
            and score_phonetique_llama >= 68.0
            and float(choix.get("score_texte") or 0.0) >= 25.0
        ):
            noyau_llama_prouve = True
            raisons_noyau_llama = [
                f"preuve_phonetique_llama_bornee={score_phonetique_llama:.1f}"
            ]
        if not noyau_llama_prouve:
            produit["produit_fiable"] = False
            produit["produit_reconnu"] = False
            produit["ambigu"] = True
            if produit.get("statut_couverture") != "HORS_COMMANDE":
                produit["statut_couverture"] = "NON_IDENTIFIE"
            produit.setdefault("raisons_ambiguite", []).append(
                "llama_rejete_sans_noyau_produit_prouve"
            )
            audits.append({
                "segment_id": produit.get("segment_id", ""),
                "applique": False,
                "raison": "llama_rejete_sans_noyau_produit_prouve",
                "declencheur": declencheur,
                "code_initial": code_initial,
                "code_article": code_choisi,
                "raisons_noyau": raisons_noyau_llama,
            })
            continue

        produit["selection"] = dict(choix)
        produit["quantite_resolue"] = choix.get("quantite_resolue")
        produit["unite_resolue"] = choix.get("unite_resolue")
        produit["produit_fiable"] = bool(
            produit.get("quantite_resolue") is not None
        )
        produit["produit_reconnu"] = produit["produit_fiable"]
        produit["ambigu"] = not produit["produit_fiable"]
        produit["statut_couverture"] = (
            "AMBIGU"
            if (
                not produit["produit_fiable"]
                or produit.get("modalite_demande") == "ALTERNATIVE"
            )
            else "RECONNU"
        )
        produit.setdefault("raisons_ambiguite", [])
        produit["raisons_ambiguite"] = [
            raison
            for raison in produit["raisons_ambiguite"]
            if raison not in {
                "selection_article_non_nette",
                "quantite_commande_non_resolue",
            }
        ]
        produit["selection"].setdefault("raisons", []).append(
            "selection_par_arbitrage_llama_produit"
        )
        produit["selection"]["regle_selection"] = (
            "arbitrage_llama_produit_borne"
        )
        audits.append({
            "segment_id": produit.get("segment_id", ""),
            "applique": True,
            "code_initial": code_initial,
            "code_article": code_choisi,
            "candidats_envoyes": len(candidats),
            "declencheur": declencheur,
            "raison": "selection_par_arbitrage_llama_produit",
        })
    return produits, audits


MOTS_OUTILS_CONSOLIDATION = {
    "a", "au", "aux", "avec", "d", "de", "des", "du", "en", "et",
    "la", "le", "les", "long", "longue", "pour", "sans", "sur", "sous",
    "dlc",
}


def _tokens_semantiques_consolidation(produit: dict[str, Any]) -> list[str]:
    """Isole les mots qui décrivent l'article, pas son conditionnement."""
    texte = str(
        produit.get("produit_normalise")
        or produit.get("texte_produit")
        or produit.get("texte_source")
        or ""
    )
    return [
        token
        for token in normaliser(texte).split()
        if token not in MOTS_OUTILS_CONSOLIDATION
        and not token.isdigit()
    ]


def _score_texte_selection(produit: dict[str, Any]) -> float:
    selection = produit.get("selection") or {}
    try:
        return float(selection.get("score_texte") or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _segments_distincts(
    produit_a: dict[str, Any],
    produit_b: dict[str, Any],
) -> bool:
    segment_a = str(produit_a.get("segment_id") or "").strip()
    segment_b = str(produit_b.get("segment_id") or "").strip()
    return bool(segment_a and segment_b and segment_a != segment_b)


def _collision_semantique_a_preserver(
    produit_a: dict[str, Any],
    produit_b: dict[str, Any],
) -> bool:
    """Détecte une collision de code, sans inférer un nouveau produit.

    Deux mentions de segments différents sont préservées quand le moteur leur
    a donné exactement la même confiance textuelle alors que le libellé forcé
    ne justifie aucune de leurs différences lexicales. C'est le signal d'un
    code intermédiaire commun, pas d'une répétition réelle.
    """
    if not _segments_distincts(produit_a, produit_b):
        return False

    tokens_a = _tokens_semantiques_consolidation(produit_a)
    tokens_b = _tokens_semantiques_consolidation(produit_b)
    if not tokens_a or not tokens_b or tokens_a == tokens_b:
        return False

    score_a = _score_texte_selection(produit_a)
    score_b = _score_texte_selection(produit_b)
    if min(score_a, score_b) < 65.0:
        return False
    if abs(score_a - score_b) > 0.01:
        return False

    selection = produit_a.get("selection") or {}
    tokens_libelle = set(
        normaliser(selection.get("libelle_article") or "").split()
    )
    differences = set(tokens_a).symmetric_difference(tokens_b)
    if not differences or differences.intersection(tokens_libelle):
        return False

    # Même famille avec deux variantes explicites (ex. base + parfum/type),
    # ou deux désignations totalement incompatibles : dans les deux cas le
    # code identique ne permet pas de conclure à une répétition.
    return tokens_a[0] == tokens_b[0] or not set(tokens_a).intersection(tokens_b)


def _cle_semantique_consolidation(produit: dict[str, Any]) -> tuple[str, ...]:
    return tuple(_tokens_semantiques_consolidation(produit))


def _fusionner_segment_ids(
    ligne: dict[str, Any],
    produit: dict[str, Any],
) -> None:
    segment_ids = [
        str(segment_id).strip()
        for segment_id in (ligne.get("segment_ids") or [])
        if str(segment_id).strip()
    ]
    segment_id = str(produit.get("segment_id") or "").strip()
    if segment_id and segment_id not in segment_ids:
        segment_ids.append(segment_id)
    if segment_ids:
        ligne["segment_ids"] = segment_ids


def construire_lignes_commande(
    produits: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    lignes_commande: list[dict[str, Any]] = []
    raisons: list[str] = []
    # L'identité d'une mention reste son segment et sa désignation. Le code
    # est seulement le résultat provisoire du matching : il peut donc porter
    # plusieurs lignes lorsqu'une collision sémantique est établie.
    index_par_article: dict[str, list[int]] = {}
    index_par_mention: dict[tuple[str, tuple[str, ...]], int] = {}
    qualite_par_ligne: dict[int, tuple[Any, ...]] = {}
    produits_par_ligne: dict[int, dict[str, Any]] = {}
    codes_avec_collision_semantique: set[str] = set()

    for index, produit in enumerate(produits, start=1):
        selection = produit.get("selection")

        if not selection:
            continue
        # Une proposition de candidat n'est pas une ligne de commande. Les
        # spans explicitement rejetes par le product gate restent dans le
        # diagnostic brut, mais ne doivent produire ni ligne ni erreur UI.
        if produit.get("produit_reconnu") is False:
            continue

        prix = selection.get("prix")
        unite_ligne = (
            charger_unites_articles().get(str(selection["code_article"]))
            or produit.get("unite_resolue")
            or produit.get("unite_principale")
        )
        if str(unite_ligne or "").upper() == "PCE":
            unite_ligne = "PI"

        quantite_ligne = (
            produit.get("quantite_resolue")
            if produit.get("quantite_resolue") is not None
            else produit.get("quantite_principale")
        )
        quantite_inferree = False
        if (
            quantite_ligne is None
            and produit.get("produit_reconnu") is True
        ):
            # Convention metier demandee : lorsqu'un vrai produit est
            # reconnu sans nombre prononce, commander une unite commerciale.
            # Cette valeur n'est jamais appliquee a un span non reconnu.
            quantite_ligne = 1.0
            quantite_inferree = True
            produit["quantite_resolue"] = 1.0
            produit["quantite_inferree"] = True
            produit["produit_fiable"] = True
            produit["statut_couverture"] = (
                "AMBIGU"
                if produit.get("modalite_demande") == "ALTERNATIVE"
                else "RECONNU"
            )
            produit["raisons_ambiguite"] = [
                raison
                for raison in produit.get("raisons_ambiguite", [])
                if raison not in {
                    "quantite_absente_a_resoudre",
                    "quantite_commande_non_resolue",
                }
            ]
            produit["ambigu"] = bool(
                produit.get("raisons_ambiguite")
            )
            raisons.append(
                f"quantite_implicite_un_ligne_{index}"
            )
        elif quantite_ligne is None:
            raisons.append(
                f"quantite_absente_ligne_{index}"
            )
            continue

        ligne = {
            "ordre_ligne": index,
            "segment_id": produit.get("segment_id", ""),
            "segment_ids": (
                [str(produit.get("segment_id")).strip()]
                if str(produit.get("segment_id") or "").strip()
                else []
            ),
            "segment_index": produit.get("segment_index", index),
            "code_article": selection[
                "code_article"
            ],
            "libelle_article": selection[
                "libelle_article"
            ],
            "quantite": quantite_ligne,
            "quantite_inferree": quantite_inferree,
            "unite": unite_ligne,
            "score_article": selection.get(
                "score_global"
            ),
            "source_recherche": selection.get(
                "source_recherche"
            ),
            "texte_source": produit.get(
                "texte_source"
            ),
            "prix": prix,
        }
        code_article = str(
            selection.get("code_article") or ""
        )
        qualite = (
            bool(produit.get("produit_fiable")),
            not bool(produit.get("ambigu")),
            float(selection.get("score_global") or 0.0),
            float(selection.get("score_selection") or 0.0),
            bool(produit.get("unite_principale")),
            index,
        )
        cle_semantique = _cle_semantique_consolidation(produit)
        cle_consolidation = (code_article, cle_semantique)
        # Sans contenu produit exploitable, il n'existe aucune preuve d'une
        # répétition : chaque segment est donc préservé.
        position = (
            index_par_mention.get(cle_consolidation)
            if cle_semantique
            else None
        )

        if position is not None:
            raisons.append(
                f"article_duplique_consolide_{code_article}"
            )
            _fusionner_segment_ids(lignes_commande[position], produit)
            if qualite > qualite_par_ligne[position]:
                lignes_commande[position] = ligne
                _fusionner_segment_ids(lignes_commande[position], produits_par_ligne[position])
                qualite_par_ligne[position] = qualite
                produits_par_ligne[position] = produit
            continue

        positions_meme_code = index_par_article.get(code_article, [])
        collision_protegee = any(
            _collision_semantique_a_preserver(
                produit,
                produits_par_ligne[position],
            )
            for position in positions_meme_code
        )
        if collision_protegee:
            codes_avec_collision_semantique.add(code_article)

        # Sans collision sémantique établie, la concurrence entre candidats
        # portant le même code conserve le meilleur résultat déjà validé par
        # le moteur. Il ne s'agit pas d'une identité de segment déduite du
        # code : les segment_id sont fusionnés seulement pour une répétition
        # sémantiquement identique, ci-dessus.
        if positions_meme_code and code_article not in codes_avec_collision_semantique:
            raisons.append(f"article_duplique_consolide_{code_article}")
            position = positions_meme_code[0]
            if qualite > qualite_par_ligne[position]:
                ancien_produit = produits_par_ligne[position]
                ancienne_cle = _cle_semantique_consolidation(ancien_produit)
                if (
                    ancienne_cle
                    and index_par_mention.get((code_article, ancienne_cle))
                    == position
                ):
                    del index_par_mention[(code_article, ancienne_cle)]
                lignes_commande[position] = ligne
                qualite_par_ligne[position] = qualite
                produits_par_ligne[position] = produit
                if cle_semantique:
                    index_par_mention[cle_consolidation] = position
            continue

        position = len(lignes_commande)
        index_par_article.setdefault(code_article, []).append(position)
        if cle_semantique:
            index_par_mention[cle_consolidation] = position
        qualite_par_ligne[position] = qualite
        lignes_commande.append(ligne)
        produits_par_ligne[position] = produit

    if not produits:
        raisons.append("produit_non_vendu_aucune_mention")
    elif not lignes_commande:
        raisons.append("produit_non_vendu_aucune_selection")

    for ordre, ligne in enumerate(
        lignes_commande,
        start=1,
    ):
        ligne["ordre_ligne"] = ordre
        produit_source = produits_par_ligne.get(ordre - 1, {})
        if not produit_source.get("produit_fiable", False):
            raisons.append(f"produit_non_fiable_ligne_{ordre}")
        if produit_source.get("ambigu", False):
            raisons.append(f"produit_ambigu_ligne_{ordre}")
        if isinstance(ligne.get("prix"), (int, float)) and ligne.get("prix") == 0:
            raisons.append(f"produit_non_vendu_prix_zero_ligne_{ordre}")
        if ligne.get("quantite") is None:
            raisons.append(f"quantite_absente_ligne_{ordre}")

    return lignes_commande, sorted(set(raisons))


def determiner_statut_commande(
    resultat: dict[str, Any],
) -> tuple[str, list[str], list[dict[str, Any]]]:
    raisons: list[str] = []

    if resultat.get("type_action_commande") == "rappel":
        return "PROBLEMATIQUE", ["message_a_rappeler"], []
    if resultat.get("type_action_commande") == "modification":
        return "PROBLEMATIQUE", ["modification_commande_a_rappeler"], []

    if not resultat.get("client_retenu"):
        raisons.extend(
            resultat.get(
                "raisons_decision_client",
                [],
            )
        )
        raisons.append(
            "client_non_mentionne_ou_non_identifie"
        )

    lignes_commande, raisons_lignes = construire_lignes_commande(
        resultat.get("produits", [])
    )
    raisons.extend(raisons_lignes)

    raisons_bloquantes = [
        r for r in raisons
        if not r.startswith("article_duplique_consolide_")
        and not r.startswith("produit_ambigu_ligne_")
        and not r.startswith("quantite_implicite_un_ligne_")
        and not (r.startswith("quantite_absente_ligne_") and lignes_commande)
        # Le prix local est une information de controle seulement. Copilote
        # applique le tarif attache au client lors de la creation : une
        # reference correctement reconnue ne doit donc pas rester a rappeler
        # uniquement parce que le cache local ne contient pas son prix.
        and not r.startswith("produit_non_vendu_prix_zero_ligne_")
    ]
    if not lignes_commande:
        raisons_bloquantes.append("produit_non_vendu_aucune_selection")

    statut = (
        "VALIDEE"
        if not raisons_bloquantes
        else "PROBLEMATIQUE"
    )

    return statut, sorted(set(raisons)), lignes_commande


def _ecrire_csv_lignes(
    chemin: Path,
    champs: list[str],
    lignes: list[dict[str, Any]],
) -> None:
    chemin.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with chemin.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as fichier:
        writer = csv.DictWriter(
            fichier,
            fieldnames=champs,
            delimiter=";",
        )
        writer.writeheader()
        writer.writerows(lignes)


def _base_export_commande(
    commande: dict[str, Any],
    run_id: str,
) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "genere_le": commande["genere_le"],
        "audio_source": commande["fichier_audio"],
        "client_code": commande.get("client_retenu"),
        "client_nom": commande.get(
            "client_nom_retenu", ""
        ),
        "date_livraison": (
            commande.get("date_livraison", {})
            or {}
        ).get("date_iso", ""),
        "statut": commande.get("statut"),
    }


def _ligne_validee_export(
    base: dict[str, Any],
    ligne: dict[str, Any],
) -> dict[str, Any]:
    return {
        **base,
        "ordre_ligne": ligne["ordre_ligne"],
        "code_article": ligne["code_article"],
        "libelle_article": ligne["libelle_article"],
        "quantite": ligne["quantite"],
        "unite": ligne["unite"],
        "score_article": ligne["score_article"],
        "source_recherche": ligne["source_recherche"],
        "texte_source": ligne["texte_source"],
        "prix": ligne["prix"],
    }


def _ligne_problematique_export(
    base: dict[str, Any],
    commande: dict[str, Any],
    raisons_supplementaires: list[str]
    | None = None,
) -> dict[str, Any]:
    raisons = list(
        commande.get("raisons_problematiques", [])
    )
    raisons.extend(raisons_supplementaires or [])

    return {
        **base,
        "statut": "PROBLEMATIQUE",
        "raisons_problematiques": " | ".join(
            sorted(set(raisons))
        ),
        "transcription": commande.get(
            "transcription", ""
        ),
        "nombre_mentions_produits": len(
            commande.get("mentions_produits", [])
        ),
    }


def _appliquer_suppression_lignes_validees(
    lignes_validees: list[dict[str, Any]],
    commande: dict[str, Any],
) -> tuple[int, list[str]]:
    client_code = commande.get("client_retenu")
    date_livraison = (
        (commande.get("date_livraison") or {}).get(
            "date_iso", ""
        )
    )
    codes_a_supprimer = Counter(
        ligne.get("code_article")
        for ligne in commande.get(
            "lignes_commande", []
        )
        if ligne.get("code_article")
    )

    if not client_code or not codes_a_supprimer:
        return 0, []

    indexes_a_supprimer: list[int] = []

    for index in range(
        len(lignes_validees) - 1,
        -1,
        -1,
    ):
        ligne = lignes_validees[index]

        if ligne.get("client_code") != client_code:
            continue

        if (
            date_livraison
            and ligne.get("date_livraison")
            != date_livraison
        ):
            continue

        code_article = ligne.get("code_article")

        if (
            code_article
            and codes_a_supprimer[code_article] > 0
        ):
            indexes_a_supprimer.append(index)
            codes_a_supprimer[code_article] -= 1

            if not any(codes_a_supprimer.values()):
                break

    for index in sorted(
        indexes_a_supprimer,
        reverse=True,
    ):
        del lignes_validees[index]

    codes_restants = [
        code_article
        for code_article, quantite_restante in (
            codes_a_supprimer.items()
        )
        for _ in range(quantite_restante)
    ]

    return len(indexes_a_supprimer), codes_restants


def preparer_exports_commandes(
    commandes: list[dict[str, Any]],
    run_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    lignes_validees: list[dict[str, Any]] = []
    lignes_problematiques: list[dict[str, Any]] = []

    for commande in commandes:
        base = _base_export_commande(
            commande=commande,
            run_id=run_id,
        )
        type_action = commande.get(
            "type_action_commande",
            "creation",
        )

        if commande.get("statut") != "VALIDEE":
            lignes_problematiques.append(
                _ligne_problematique_export(
                    base=base,
                    commande=commande,
                )
            )
            continue

        if type_action == "suppression":
            nb_supprimees, codes_restants = (
                _appliquer_suppression_lignes_validees(
                    lignes_validees=lignes_validees,
                    commande=commande,
                )
            )

            if nb_supprimees == 0:
                lignes_problematiques.append(
                    _ligne_problematique_export(
                        base=base,
                        commande=commande,
                        raisons_supplementaires=[
                            "suppression_produit_introuvable"
                        ],
                    )
                )
                continue

            if codes_restants:
                raisons = [
                    "suppression_partielle",
                    "codes_restants="
                    + ",".join(codes_restants),
                ]
                lignes_problematiques.append(
                    _ligne_problematique_export(
                        base=base,
                        commande=commande,
                        raisons_supplementaires=raisons,
                    )
                )

            continue

        for ligne in commande.get(
            "lignes_commande", []
        ):
            lignes_validees.append(
                _ligne_validee_export(
                    base=base,
                    ligne=ligne,
                )
            )

    return lignes_validees, lignes_problematiques


def exporter_csv_commandes(
    commandes: list[dict[str, Any]],
    run_id: str,
) -> tuple[Path, Path]:
    (
        lignes_validees,
        lignes_problematiques,
    ) = preparer_exports_commandes(
        commandes=commandes,
        run_id=run_id,
    )

    chemin_validees = (
        DOSSIER_COMMANDES_VALIDEES
        / "commandes_validees.csv"
    )
    chemin_problematiques = (
        DOSSIER_COMMANDES_PROBLEMATIQUES
        / "commandes_problematiques.csv"
    )

    _ecrire_csv_lignes(
        chemin=chemin_validees,
        champs=[
            "run_id",
            "genere_le",
            "audio_source",
            "client_code",
            "client_nom",
            "date_livraison",
            "statut",
            "ordre_ligne",
            "code_article",
            "libelle_article",
            "quantite",
            "unite",
            "score_article",
            "source_recherche",
            "texte_source",
            "prix",
        ],
        lignes=lignes_validees,
    )

    _ecrire_csv_lignes(
        chemin=chemin_problematiques,
        champs=[
            "run_id",
            "genere_le",
            "audio_source",
            "client_code",
            "client_nom",
            "date_livraison",
            "statut",
            "raisons_problematiques",
            "nombre_mentions_produits",
            "transcription",
        ],
        lignes=lignes_problematiques,
    )

    return chemin_validees, chemin_problematiques


# -------------------------------------------------------------------
# Programme principal
# -------------------------------------------------------------------

DATE_FICHIER_REPONDEUR_RE = re.compile(r"(\d{4}-\d{2}-\d{2})[_ -](\d{2})[-h](\d{2})")


def date_reference_depuis_nom_fichier(nom_fichier: str) -> date | None:
    match = DATE_FICHIER_REPONDEUR_RE.search(nom_fichier)
    if not match:
        return None
    try:
        return date.fromisoformat(match.group(1))
    except ValueError:
        return None


def heure_reference_depuis_nom_fichier(nom_fichier: str) -> int | None:
    match = DATE_FICHIER_REPONDEUR_RE.search(nom_fichier)
    if not match:
        return None
    try:
        return int(match.group(2))
    except ValueError:
        return None


def telephone_depuis_nom_fichier(nom_fichier: str) -> str:
    match = re.search(
        r"_De-([^_.]+)",
        nom_fichier,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""
    return normaliser_telephone(match.group(1))


def analyser_structure_commande(transcription: str) -> dict[str, str]:
    """
    Sépare la transcription brute en composants sémantiques (client, date, produits)
    sans supposer un ordre d'apparition ni une continuité.
    """
    t = re.sub(r'\b(et\s+demie?)\b', lambda m: m.group(1).replace(' ', '_'), transcription.lower().strip())
    
    # Split on explicit pauses/punctuation
    morceaux = re.split(r'[.,;!]+', t)
    
    # Split on " et " if it seems to link clauses (not inside a product like "sel et poivre")
    clauses = []
    for m in morceaux:
        sub = re.split(r'\bet\s+(?=(?:on|je|il|ça|c|ce|pour)\b)', m)
        clauses.extend([s.strip() for s in sub if s.strip()])
        
    client_parts = []
    date_parts = []
    produit_parts = []
    
    mots_date = r'\b(pour|demain|lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche|matin|soir|aujourd hui|livraison)\b'
    mots_client = r'\b(client|restaurant|compte|part\s+de|nom\s+de|ici|c\s+est)\b'
    motif_qte = r'\b(\d+(?:\.\d+)?|un|une|deux|trois|quatre|cinq|six|sept|huit|neuf|dix|douze|vingt|trente|quarante|cinquante|soixante|cent)\b'
    
    for clause in clauses:
        c_norm = clause.replace('_', ' ')
        
        has_qte = bool(re.search(motif_qte, c_norm))
        has_date = bool(re.search(mots_date, c_norm))
        has_client = bool(re.search(mots_client, c_norm))
        
        if has_date and not re.search(r'(litre|kilo|boite|carton|poche|seau|bidon|pack|colis|gramme)', c_norm):
            date_parts.append(c_norm)
        elif has_client and not has_qte:
            client_parts.append(c_norm)
        elif has_qte:
            if re.search(r'\b(0\d(\s*\d{2}){4})\b', c_norm):
                client_parts.append(c_norm)
            else:
                produit_parts.append(c_norm)
        else:
            # S'il n'y a pas de quantité mais que ça contient un mot de produit ou c'est orphelin
            # On l'ajoute aux produits pour éviter de perdre des infos comme "surgelé", etc.
            produit_parts.append(c_norm)
            
    return {
        "client_brut": " , ".join(client_parts),
        "date_brute": " , ".join(date_parts),
        "produits_bruts": " , ".join(produit_parts)
    }


def traiter_transcriptions(
    chemins_transcriptions: list[Path] | None = None,
    date_reference: date | None = None,
) -> list[dict[str, Any]]:
    DOSSIER_RESULTATS.mkdir(
        parents=True,
        exist_ok=True,
    )

    clients_bruts = charger_clients()
    cadencier_brut = charger_cadencier()

    # Règle métier : un client est actif/valide uniquement s'il est présent à la fois
    # dans cadencier-clientsBASCO ET dans info-clients.
    codes_info = {c["code_client"] for c in clients_bruts}
    codes_cadencier = set(cadencier_brut.keys())
    codes_actifs = codes_info.intersection(codes_cadencier)

    # Exclure de la reconnaissance client tout client absent de l'un des deux fichiers
    clients = [c for c in clients_bruts if c["code_client"] in codes_actifs]
    cadencier = {k: v for k, v in cadencier_brut.items() if k in codes_actifs}

    unites_articles = charger_unites_articles()
    for produits_client in cadencier.values():
        for produit in produits_client:
            produit["unite_vente"] = unites_articles.get(
                str(produit.get("code_article") or ""), ""
            )
    articles_reference = charger_catalogue_articles_reference()
    for article in articles_reference:
        code = str(article.get("code_article") or "")
        article["unite_vente"] = unites_articles.get(code, "")

    nb_clients_ajoutes = enrichir_clients_depuis_cadencier(
        clients=clients,
        cadencier=cadencier,
    )
    stats_ventes = charger_stats_ventes_clients()
    enrichir_clients_avec_stats_ventes(
        clients=clients,
        stats_ventes=stats_ventes,
    )
    catalogue_global = construire_catalogue_global(
        cadencier,
        articles_reference=articles_reference,
    )
    catalogue_reappro = charger_reappro_basco()
    variantes_clients = charger_variantes_clients(
        CHEMIN_VARIANTES_CLIENTS
    )
    telephones_clients = charger_telephones_clients(
        CHEMIN_TELEPHONES_CLIENTS
    )
    aliases_telephoniques_confirmes = (
        charger_aliases_telephoniques_confirmes(
            CHEMIN_ALIASES_TELEPHONIQUES_CONFIRMES
        )
    )
    synonymes_produits = charger_synonymes_produits(
        CHEMIN_SYNONYMES_PRODUITS
    )
    enrichir_alias_avec_variantes(
        clients=clients,
        variantes_par_code=variantes_clients,
    )
    enrichir_clients_avec_telephones(
        clients=clients,
        telephones_par_code=telephones_clients,
    )
    enrichir_clients_avec_aliases_telephoniques_confirmes(
        clients=clients,
        aliases_confirmes=aliases_telephoniques_confirmes,
    )

    fichiers = (
        sorted(chemins_transcriptions)
        if chemins_transcriptions is not None
        else sorted(
            DOSSIER_TRANSCRIPTIONS.glob(
                "*__transcription.json"
            )
        )
    )

    if not fichiers:
        raise RuntimeError(
            "Aucune transcription JSON trouvée dans : "
            f"{DOSSIER_TRANSCRIPTIONS}"
        )

    print(f"Clients chargés : {len(clients)}")
    print(
        "Clients ajoutés depuis cadencier : "
        f"{nb_clients_ajoutes}"
    )
    print(f"Clients avec cadencier : {len(cadencier)}")
    print(
        "Catalogue global articles : "
        f"{len(catalogue_global)}"
    )
    print(
        "Clients avec variantes manuelles : "
        f"{len(variantes_clients)}"
    )
    print(
        "Clients avec telephones manuels : "
        f"{len(telephones_clients)}"
    )
    print(
        "Entrées synonymes produits : "
        f"{len(synonymes_produits)}"
    )
    print("")

    commandes: list[dict[str, Any]] = []

    for chemin in fichiers:
        transcription = lire_transcription(
            chemin
        )
        date_reference_commande = (
            date_reference
            if date_reference is not None
            else date_reference_depuis_nom_fichier(chemin.name)
        )
        action_commande = detecter_type_action_commande(
            transcription
        )

        mentions = (
            []
            if action_commande["type_action"] in {"rappel", "modification"}
            else extraire_mentions_produits(transcription)
        )
        if (
            action_commande["type_action"]
            == "suppression"
            and not mentions
        ):
            mentions = extraire_mentions_suppression(
                transcription
            )

        # Identifiant cree avant toute recherche, filtre ou consolidation :
        # il relie une mention a sa ligne de commande sans position implicite.
        for segment_index, mention in enumerate(mentions, start=1):
            mention["segment_id"] = f"segment-{segment_index}"
            mention["segment_index"] = segment_index

        identification_client = identifier_client(
            transcription=transcription,
            clients=clients,
            cadencier=cadencier,
            mentions_produits=mentions,
            telephone_appel=telephone_depuis_nom_fichier(
                chemin.name
            ),
        )

        # Arbitrage local borne : le LLM ne voit que les clients actifs deja
        # preselctionnes par le moteur deterministe et ne peut en inventer
        # aucun. Il est reserve aux conflits de preuves client.
        arbitrage_client_requis = client_requiert_arbitrage_llm(
            identification_client.get("candidats", []),
            identification_client.get("client_retenu"),
        )
        identification_client["arbitrage_llm_client"] = {
            "requis": arbitrage_client_requis,
            "applique": False,
            "candidats_envoyes": 0,
            "raison": "",
        }
        if arbitrage_client_requis:
            candidats_llm = candidats_pour_arbitrage_llm(
                identification_client.get("candidats", []),
                identification_client.get("client_retenu"),
                limite=15,
            )
            identification_client["arbitrage_llm_client"][
                "candidats_envoyes"
            ] = len(candidats_llm)
            try:
                from src.llm_arbitrage import (
                    arbitrer_client_ambigu,
                    ollama_disponible,
                )

                choix_llm = (
                    arbitrer_client_ambigu(
                        zone_client=identification_client.get(
                            "zone_client", ""
                        ),
                        candidats=candidats_llm,
                    )
                    if ollama_disponible()
                    else None
                )
            except Exception:
                choix_llm = None

            if choix_llm:
                identification_client["client_retenu"] = choix_llm.get(
                    "code_client"
                )
                identification_client["client_nom_retenu"] = choix_llm.get(
                    "nom_client", ""
                )
                identification_client["decision_automatique"] = True
                identification_client["raisons_decision"] = [
                    "client_identifie_par_arbitrage_llama_actifs",
                ]
                identification_client["arbitrage_llm_client"].update(
                    {
                        "applique": True,
                        "raison": "conflit_preuves_client",
                        "code_client": choix_llm.get("code_client", ""),
                    }
                )
            else:
                identification_client["arbitrage_llm_client"][
                    "raison"
                ] = "llama_indisponible_ou_sans_choix"

        clients_candidats = identification_client[
            "candidats"
        ]

        client_retenu = identification_client[
            "client_retenu"
        ]
        client_nom_retenu = identification_client.get(
            "client_nom_retenu", ""
        )

        client_actif_retenu = next(
            (
                client
                for client in clients
                if client.get("code_client") == client_retenu
            ),
            None,
        )
        candidat_client_retenu = next(
            (
                candidat
                for candidat in clients_candidats
                if candidat.get("code_client") == client_retenu
            ),
            None,
        )
        mentions, mentions_client_exclues = filtrer_mentions_client_resolu(
            mentions=mentions,
            client=client_actif_retenu,
            candidat_retenu=candidat_client_retenu,
            zone_client=identification_client.get("zone_client", ""),
        )
        identification_client["mentions_client_exclues"] = (
            mentions_client_exclues
        )

        produits = (
            []
            if action_commande["type_action"] in {"rappel", "modification"}
            else chercher_produits(
                mentions=mentions,
                produits_client=cadencier.get(
                    client_retenu,
                    [],
                ),
                catalogue_global=catalogue_global,
                synonymes_produits=synonymes_produits,
                catalogue_reappro=catalogue_reappro,
            )
        )
        produits, arbitrages_llm_produits = (
            arbitrer_produits_ambigus_llama(
                produits=produits,
                client_nom=client_nom_retenu,
            )
        )
        produits, commande_courte_configuree = (
            appliquer_commande_courte_client(
                transcription=transcription,
                client_code=client_retenu,
                type_action=action_commande["type_action"],
                mentions=mentions,
                produits=produits,
                produits_client=cadencier.get(client_retenu, []),
            )
        )

        resultat = {
            "fichier_audio": chemin.name.replace(
                "__transcription.json",
                ".ogg",
            ),
            "fichier_transcription": chemin.name,
            "genere_le": datetime.now().isoformat(),
            "transcription": transcription,
            "type_action_commande": action_commande[
                "type_action"
            ],
            "expression_action_commande": action_commande[
                "expression"
            ],
            "zone_client_detectee": identification_client[
                "zone_client"
            ],
            "clients_candidats": clients_candidats,
            "client_retenu": client_retenu,
            "client_nom_retenu": client_nom_retenu,
            "decision_automatique_client": identification_client[
                "decision_automatique"
            ],
            "raisons_decision_client": identification_client[
                "raisons_decision"
            ],
            "identification_client": identification_client,
            "date_livraison": resoudre_date_livraison(
                transcription,
                date_reference=date_reference_commande,
                heure_reference=heure_reference_depuis_nom_fichier(
                    chemin.name
                ),
            ),
            "mentions_produits": mentions,
            "produits": produits,
            "arbitrages_llm_produits": arbitrages_llm_produits,
            "commande_courte_configuree": commande_courte_configuree,
        }

        statut, raisons_problematiques, lignes_commande = (
            determiner_statut_commande(
                resultat
            )
        )
        resultat["statut"] = statut
        resultat[
            "raisons_problematiques"
        ] = raisons_problematiques
        resultat[
            "lignes_commande"
        ] = lignes_commande

        resultat["ai_arbitrage"] = {
            "enabled": False,
            "api_key_available": False,
            "applied": False,
            "skipped": "traitement_exclusivement_sur_instance_locale",
        }

        chemin_json = (
            DOSSIER_RESULTATS
            / chemin.name.replace(
                "__transcription.json",
                "__extraction.json",
            )
        )

        chemin_txt = (
            DOSSIER_RESULTATS
            / chemin.name.replace(
                "__transcription.json",
                "__extraction.txt",
            )
        )

        chemin_json.write_text(
            json.dumps(
                resultat,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        chemin_txt.write_text(
            creer_resume_txt(resultat),
            encoding="utf-8",
        )

        print(f"Extraction : {chemin.name}")
        print(
            "  action : "
            f"{action_commande['type_action']}"
        )
        print(
            "  client retenu : "
            f"{client_retenu or 'aucun - à vérifier'}"
        )
        print(
            "  décision auto client : "
            + (
                "OUI"
                if identification_client[
                    "decision_automatique"
                ]
                else "NON"
            )
        )
        print(f"  mentions produits : {len(mentions)}")
        print(f"  statut : {statut}")
        print(f"  TXT : {chemin_txt}")
        print("")

        commandes.append(resultat)

    return commandes


def main() -> None:
    run_id = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )
    commandes = traiter_transcriptions()
    chemin_validees, chemin_problematiques = (
        exporter_csv_commandes(
            commandes=commandes,
            run_id=run_id,
        )
    )
    print("CSV commandes validées :", chemin_validees)
    print(
        "CSV commandes problématiques :",
        chemin_problematiques,
    )


if __name__ == "__main__":
    main()
