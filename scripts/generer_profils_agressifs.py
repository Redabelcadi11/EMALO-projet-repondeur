import sys, json, pathlib
from openpyxl import load_workbook
from collections import defaultdict

EXCEL = pathlib.Path('resultats/evaluation-copilote/comparaison_ES_vs_logiciel.xlsx')
CONFIG = pathlib.Path('config/profils-clients-agressifs.json')

def run():
    wb = load_workbook(EXCEL)
    ws = wb[wb.sheetnames[1]]
    
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
        
    commandes = []
    current = None
    for row in rows:
        if row.get("N°"):
            if current: commandes.append(current)
            current = {
                "client": row.get("Client ES (code)", str(row.get("Client Logiciel", ""))).strip(),
                "articles": []
            }
        if current and row.get("Code Article"):
            current["articles"].append({
                "code": str(row["Code Article"]).strip(),
                "source": row.get("Source", "")
            })
    if current: commandes.append(current)
    
    # Analysis
    subs_count = defaultdict(lambda: defaultdict(int))
    ghost_count = defaultdict(lambda: defaultdict(int))
    total_orders = defaultdict(int)
    
    for cmd in commandes:
        client = cmd["client"]
        if not client: continue
        total_orders[client] += 1
        
        fp = []
        miss = []
        for a in cmd["articles"]:
            if a["source"] == "Logiciel seulement": fp.append(a["code"])
            elif a["source"] == "ES seulement": miss.append(a["code"])
            
        if len(fp) >= 1 and len(miss) >= 1:
            for f in fp:
                for m in miss:
                    subs_count[client][(f, m)] += 1
            
        for m in miss:
            ghost_count[client][m] += 1
            
    print(f"Total commands analyzed: {sum(total_orders.values())} for {len(total_orders)} clients")
    print(f"Substitutions candidates: {len(subs_count)} clients have candidates")
    for c, pairs in subs_count.items():
        print(f"  Client {c}: {len(pairs)} pair(s) - e.g. {list(pairs.items())[:2]}")
            
    agressif = {"substitutions": {}, "ajouts_fantomes": {}}
    
    # Generate Substitutions (>= 1 times)
    for client, pairs in subs_count.items():
        c_subs = {}
        for (f, m), count in pairs.items():
            if count >= 1: c_subs[f] = m
        if c_subs: agressif["substitutions"][client] = c_subs
        
    # Generate Ghosts (>= 1 times AND >= 50% of orders)
    for client, items in ghost_count.items():
        c_ghosts = []
        tot = total_orders[client]
        for m, count in items.items():
            if count >= 1 and count / tot >= 0.5:
                c_ghosts.append(m)
        if c_ghosts: agressif["ajouts_fantomes"][client] = c_ghosts
        
    CONFIG.parent.mkdir(exist_ok=True)
    CONFIG.write_text(json.dumps(agressif, indent=4), encoding='utf-8')
    print("PROFILES GENERES:")
    print(json.dumps(agressif, indent=2))
    print(f"\nSauvegardé dans: {CONFIG}")

if __name__ == '__main__':
    run()
