"""Transcrit des echantillons d'appels pour identifier manuellement les clients.

Ce script ne lance ni extraction de commande, ni appel ERP. Il selectionne au
plus N audios recents par numero absent de la base telephonique et ecrit
uniquement les transcriptions locales produites par le worker GPU.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_ROOT = Path(__file__).resolve().parents[1]
if str(SCRIPT_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPT_ROOT))

from prod_pipeline import all_nextcloud_audios, ensure_transcriptions_for_audios, transcription_path_for
from src.clients import (
    charger_aliases_telephoniques_confirmes,
    charger_telephones_clients,
    normaliser_telephone,
    normaliser_telephones,
)
from src.runtime_paths import bootstrap_runtime_environment, get_project_root


PROJECT_ROOT = bootstrap_runtime_environment()
INFO_CLIENTS_PATH = (
    PROJECT_ROOT
    / "ressources-originales"
    / "informations-clients"
    / "info-clients.xlsx"
)
PHONE_CONFIG_PATH = PROJECT_ROOT / "config" / "telephones-clients.json"
CONFIRMED_ALIASES_PATH = PROJECT_ROOT / "config" / "aliases-telephoniques-confirmes.json"


def phones_from_info_clients(path: Path = INFO_CLIENTS_PATH) -> set[str]:
    """Read only the real workbook, never Excel's temporary ~$ lock file."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        phone_index = next(
            index for index, header in enumerate(headers) if "phone" in header.lower()
        )
        phones: set[str] = set()
        for row in rows:
            if phone_index < len(row):
                phones.update(normaliser_telephones(row[phone_index]))
        return phones
    finally:
        workbook.close()


def caller_phone(audio_name: str) -> str:
    from generer_ui_data_prod import parse_phone

    return normaliser_telephone(parse_phone(audio_name))


def unknown_phone_samples(max_per_phone: int) -> dict[str, list[Path]]:
    known_info = phones_from_info_clients()
    known_config = {
        phone
        for phones in charger_telephones_clients(PHONE_CONFIG_PATH).values()
        for phone in phones
    }
    known_aliases = set(
        charger_aliases_telephoniques_confirmes(CONFIRMED_ALIASES_PATH)
    )
    grouped: dict[str, list[Path]] = defaultdict(list)
    for audio in all_nextcloud_audios():
        phone = caller_phone(audio.name)
        if not phone or phone in known_info or phone in known_config or phone in known_aliases:
            continue
        grouped[phone].append(audio)

    return {
        phone: audios[:max_per_phone]
        for phone, audios in sorted(
            grouped.items(), key=lambda item: (-len(item[1]), item[0])
        )
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-per-phone", type=int, default=10)
    parser.add_argument(
        "--limit-phones",
        type=int,
        default=0,
        help="0 = tous les numeros absents de la base",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)
    if args.max_per_phone < 1:
        parser.error("--max-per-phone doit etre superieur ou egal a 1")

    groups = unknown_phone_samples(args.max_per_phone)
    if args.limit_phones > 0:
        selected_phones = list(groups)[: args.limit_phones]
        groups = {phone: groups[phone] for phone in selected_phones}

    selected = [
        (phone, audio)
        for phone, audios in groups.items()
        for audio in audios
    ]
    missing = [audio for _, audio in selected if not transcription_path_for(audio).exists()]
    summary = {
        "phones": len(groups),
        "audios_selected": len(selected),
        "already_transcribed": len(selected) - len(missing),
        "to_transcribe": len(missing),
        "max_per_phone": args.max_per_phone,
        "mode": "transcription_only",
    }
    print(json.dumps({"event": "start", **summary}, ensure_ascii=False), flush=True)
    if args.dry_run:
        return 0

    completed = 0
    failed: list[dict[str, str]] = []
    for position, (phone, audio) in enumerate(selected, start=1):
        if transcription_path_for(audio).exists():
            completed += 1
            continue
        try:
            ensure_transcriptions_for_audios([audio])
            completed += 1
            print(
                json.dumps(
                    {
                        "event": "progress",
                        "position": position,
                        "total": len(selected),
                        "phone": phone,
                        "audio": audio.name,
                        "completed": completed,
                        "failed": len(failed),
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )
        except Exception as exc:
            failed.append({"phone": phone, "audio": audio.name, "error": str(exc)})
            print(
                json.dumps(
                    {
                        "event": "error",
                        "position": position,
                        "phone": phone,
                        "audio": audio.name,
                        "error": str(exc),
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )

    print(
        json.dumps(
            {
                "event": "done",
                **summary,
                "completed": completed,
                "failed": len(failed),
                "failures": failed[:20],
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
