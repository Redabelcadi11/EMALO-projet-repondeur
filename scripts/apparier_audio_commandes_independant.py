#!/usr/bin/env python3
"""Appariement prive audio/commande, independant des predictions EMALO."""
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz


GENERIC_CLIENT_TOKENS = {
    "bar", "chez", "client", "hotel", "la", "le", "les", "restaurant",
    "sarl", "sas", "sasu", "societe", "snack",
}
PRODUCT_STOPWORDS = {
    "avec", "barquette", "barquettes", "bidon", "bidons", "boite", "boites",
    "carton", "cartons", "colis", "dans", "des", "deux", "gramme", "grammes",
    "kilo", "kilos", "litre", "litres", "pack", "paquet", "paquets", "piece",
    "pieces", "poche", "poches", "pot", "pots", "pour", "sac", "sacs",
    "sachet", "sachets", "seau", "seaux", "une",
}


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(re.findall(r"[a-z0-9]+", text))


def normalize_phone(value: Any) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    if digits.startswith("0033") and len(digits) >= 12:
        return "0" + digits[4:]
    if digits.startswith("33") and len(digits) >= 11:
        return "0" + digits[2:]
    return digits


def phone_from_audio(name: str) -> str:
    match = re.search(r"_De-([^_.]+)", name, re.IGNORECASE)
    return normalize_phone(match.group(1)) if match else ""


def date_from_audio(name: str) -> str:
    try:
        return date.fromisoformat(name[:10]).isoformat()
    except ValueError:
        return ""


def content_tokens(value: Any, stopwords: set[str]) -> list[str]:
    return [
        token for token in normalize(value).split()
        if len(token) >= 3 and not token.isdigit() and token not in stopwords
    ]


def atomic_json(path: Path, payload: dict[str, Any], mode: int = 0o600) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=path.name + ".", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(handle, "w", encoding="utf-8") as output:
            json.dump(payload, output, ensure_ascii=False, indent=2)
            output.write("\n")
        os.chmod(temporary_name, mode)
        os.replace(temporary_name, path)
    finally:
        if os.path.exists(temporary_name):
            os.unlink(temporary_name)


def load_clients(
    workbook_path: Path,
    manual_phones_path: Path | None = None,
    variants_path: Path | None = None,
) -> tuple[dict[str, dict[str, Any]], dict[str, set[str]]]:
    clients: dict[str, dict[str, Any]] = {}
    phones: dict[str, set[str]] = defaultdict(set)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    next(iterator, None)
    for row in iterator:
        code = str(row[2] or "").strip().upper()
        if not code:
            continue
        profile = clients.setdefault(
            code,
            {
                "code": code,
                "labels": set(),
                "cities": set(),
                "aliases": set(),
            },
        )
        label = str(row[3] or "").strip()
        city = str(row[12] or "").strip()
        if label:
            profile["labels"].add(label)
        if city:
            profile["cities"].add(city)
        phone = normalize_phone(row[13])
        if phone:
            phones[phone].add(code)
    workbook.close()

    if variants_path and variants_path.is_file():
        variants = json.loads(variants_path.read_text(encoding="utf-8"))
        for code, values in variants.items():
            profile = clients.setdefault(
                str(code).upper(),
                {"code": str(code).upper(), "labels": set(), "cities": set(), "aliases": set()},
            )
            profile["aliases"].update(str(value) for value in values if str(value).strip())
    if manual_phones_path and manual_phones_path.is_file():
        manual = json.loads(manual_phones_path.read_text(encoding="utf-8"))
        for code, values in manual.items():
            for value in values:
                phone = normalize_phone(value)
                if phone:
                    phones[phone].add(str(code).upper())
    return clients, phones


def load_orders(csv_paths: list[Path]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for path in csv_paths:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle, delimiter=";"):
                if str(row.get("operator") or "").strip().upper() != "ES":
                    continue
                number = str(row.get("order_number") or "").strip()
                if number:
                    grouped[number].append(dict(row))
    orders: list[dict[str, Any]] = []
    for number, rows in grouped.items():
        first = rows[0]
        lines = [
            {
                "code": str(row.get("article_code") or "").strip(),
                "label": str(row.get("designation") or "").strip(),
                "quantity": str(row.get("quantity") or "").strip(),
                "unit": str(row.get("unit") or "").strip().upper(),
            }
            for row in rows
            if str(row.get("article_code") or "").strip()
        ]
        orders.append(
            {
                "order_number": number,
                "client_code": str(first.get("client_code") or "").strip().upper(),
                "client_label": str(first.get("client_label") or "").strip(),
                "order_date": str(first.get("order_date") or "")[:10],
                "delivery_date": str(
                    first.get("delivery_date") or first.get("departure_date") or ""
                )[:10],
                "lines": lines,
                "has_error": any(str(row.get("error") or "").strip() for row in rows),
            }
        )
    return sorted(orders, key=lambda row: (row["order_date"], row["order_number"]))


def load_audios(transcriptions: Path, date_from: str, date_to: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in sorted(transcriptions.glob("*__transcription.json")):
        audio_stem = path.name.removesuffix("__transcription.json")
        audio_date = date_from_audio(audio_stem)
        if not audio_date or not date_from <= audio_date <= date_to:
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        audio = str(payload.get("fichier_audio") or audio_stem)
        rows.append(
            {
                "audio": audio,
                "date": audio_date,
                "phone": phone_from_audio(audio),
                "transcription": str(payload.get("texte") or "").strip(),
                "transcription_sha256": __import__("hashlib").sha256(path.read_bytes()).hexdigest(),
            }
        )
    return rows


def client_text_score(
    transcription: str,
    order: dict[str, Any],
    profile: dict[str, Any] | None,
) -> tuple[float, list[str]]:
    values = [order.get("client_label") or ""]
    cities: list[str] = []
    if profile:
        values.extend(profile.get("labels") or [])
        values.extend(profile.get("aliases") or [])
        cities.extend(profile.get("cities") or [])
    text = normalize(transcription)
    text_tokens = set(content_tokens(text, GENERIC_CLIENT_TOKENS))
    best = 0.0
    evidence: list[str] = []
    for value in values:
        tokens = content_tokens(value, GENERIC_CLIENT_TOKENS)
        if not tokens:
            continue
        phrase = " ".join(tokens)
        exact = sum(token in text_tokens for token in tokens) / len(tokens)
        fuzzy = fuzz.partial_ratio(phrase, text) / 100.0
        score = 0.65 * exact + 0.35 * fuzzy
        if phrase in text:
            score = max(score, 0.95)
        if score > best:
            best = score
            evidence = [phrase]
    city_score = 0.0
    for city in cities:
        normalized_city = normalize(city)
        if normalized_city and normalized_city in text:
            city_score = 1.0
            evidence.append(f"ville:{normalized_city}")
            break
    return min(1.0, 0.82 * best + 0.18 * city_score), evidence


def product_line_score(label: str, transcription: str) -> float:
    label_tokens = content_tokens(label, PRODUCT_STOPWORDS)
    text_tokens = content_tokens(transcription, PRODUCT_STOPWORDS)
    if not label_tokens or not text_tokens:
        return 0.0
    exact = sum(token in text_tokens for token in label_tokens) / len(label_tokens)
    fuzzy = sum(
        max(fuzz.ratio(token, heard) for heard in text_tokens) / 100.0
        for token in label_tokens
    ) / len(label_tokens)
    partial = fuzz.partial_ratio(" ".join(label_tokens), normalize(transcription)) / 100.0
    return 0.35 * exact + 0.30 * fuzzy + 0.35 * partial


def product_evidence(order: dict[str, Any], transcription: str) -> dict[str, Any]:
    scores = [product_line_score(line["label"], transcription) for line in order["lines"]]
    if not scores:
        return {"score": 0.0, "coverage": 0.0, "matches": 0, "best": 0.0}
    matches = sum(score >= 0.48 for score in scores)
    coverage = matches / len(scores)
    best_scores = sorted(scores, reverse=True)[: min(3, len(scores))]
    top_mean = sum(best_scores) / len(best_scores)
    evidence = sum(max(0.0, score - 0.30) for score in scores) / math.sqrt(len(scores))
    score = 0.50 * coverage + 0.30 * top_mean + 0.20 * min(1.0, evidence)
    return {
        "score": round(score, 6),
        "coverage": round(coverage, 6),
        "matches": matches,
        "best": round(max(scores), 6),
    }


def build_edges(
    audios: list[dict[str, Any]],
    orders: list[dict[str, Any]],
    clients: dict[str, dict[str, Any]],
    phones: dict[str, set[str]],
) -> list[dict[str, Any]]:
    orders_by_date: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for order in orders:
        if order["lines"] and not order["has_error"]:
            orders_by_date[order["order_date"]].append(order)
    edges: list[dict[str, Any]] = []
    for audio in audios:
        phone_clients = phones.get(audio["phone"], set())
        for order in orders_by_date.get(audio["date"], []):
            code = order["client_code"]
            phone_unique = len(phone_clients) == 1 and code in phone_clients
            phone_ambiguous = len(phone_clients) > 1 and code in phone_clients
            phone_conflict = bool(phone_clients and code not in phone_clients)
            text_score, client_evidence = client_text_score(
                audio["transcription"], order, clients.get(code)
            )
            client_score = 1.0 if phone_unique else 0.86 if phone_ambiguous else text_score
            products = product_evidence(order, audio["transcription"])
            if phone_unique:
                total = 0.60 + 0.40 * products["score"]
            else:
                total = 0.52 * client_score + 0.48 * products["score"]
            if phone_conflict:
                total *= 0.15
            edges.append(
                {
                    "audio": audio["audio"],
                    "order_number": order["order_number"],
                    "date": audio["date"],
                    "client_code": code,
                    "score": round(total, 6),
                    "client_score": round(client_score, 6),
                    "phone_unique": phone_unique,
                    "phone_ambiguous": phone_ambiguous,
                    "phone_conflict": phone_conflict,
                    "client_evidence": client_evidence,
                    "product_score": products["score"],
                    "product_coverage": products["coverage"],
                    "product_matches": products["matches"],
                    "product_best": products["best"],
                }
            )
    return edges


def rank_edges(edges: list[dict[str, Any]]) -> None:
    by_audio: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_order: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for edge in edges:
        by_audio[edge["audio"]].append(edge)
        by_order[edge["order_number"]].append(edge)
    for items, rank_key, margin_key in (
        (by_audio, "audio_rank", "audio_margin"),
        (by_order, "order_rank", "order_margin"),
    ):
        for candidates in items.values():
            candidates.sort(key=lambda item: (-item["score"], item["audio"], item["order_number"]))
            for index, candidate in enumerate(candidates):
                candidate[rank_key] = index + 1
                next_score = candidates[index + 1]["score"] if index + 1 < len(candidates) else 0.0
                candidate[margin_key] = round(candidate["score"] - next_score, 6)


def classify_edges(
    edges: list[dict[str, Any]],
    audios: list[dict[str, Any]],
    orders: list[dict[str, Any]],
    phones: dict[str, set[str]],
) -> None:
    audio_by_name = {audio["audio"]: audio for audio in audios}
    audio_groups: Counter[tuple[str, str]] = Counter()
    for audio in audios:
        mapped = phones.get(audio["phone"], set())
        if len(mapped) == 1:
            audio_groups[(audio["date"], next(iter(mapped)))] += 1
    order_groups = Counter(
        (order["order_date"], order["client_code"])
        for order in orders
        if order["lines"] and not order["has_error"]
    )
    for edge in edges:
        group = (edge["date"], edge["client_code"])
        metadata_exact = bool(
            edge["phone_unique"]
            and audio_groups[group] == 1
            and order_groups[group] == 1
            and edge.get("audio_rank") == 1
            and edge.get("order_rank") == 1
            and edge.get("order_margin", 0.0) >= 0.08
            and (
                edge["product_best"] >= 0.52
                or edge["product_coverage"] >= 0.20
                or edge["client_score"] >= 0.95
            )
        )
        semantic_high = bool(
            not edge["phone_conflict"]
            and edge.get("audio_rank") == 1
            and edge.get("order_rank") == 1
            and edge["score"] >= 0.62
            and edge["client_score"] >= 0.55
            and edge["audio_margin"] >= 0.10
            and edge["order_margin"] >= 0.08
            and (
                edge["product_coverage"] >= 0.34
                or edge["product_matches"] >= 2
            )
        )
        edge["confidence_class"] = (
            "metadata_exact" if metadata_exact else "semantic_high" if semantic_high else "rejected"
        )
        edge["audio_phone"] = audio_by_name[edge["audio"]]["phone"]


def select_assignments(edges: list[dict[str, Any]]) -> list[dict[str, Any]]:
    priority = {"metadata_exact": 0, "semantic_high": 1}
    candidates = [edge for edge in edges if edge["confidence_class"] in priority]
    candidates.sort(
        key=lambda edge: (
            priority[edge["confidence_class"]],
            -edge["score"],
            -edge["audio_margin"],
            -edge["order_margin"],
            edge["audio"],
        )
    )
    used_audio: set[str] = set()
    used_order: set[str] = set()
    selected: list[dict[str, Any]] = []
    for edge in candidates:
        if edge["audio"] in used_audio or edge["order_number"] in used_order:
            continue
        selected.append(edge)
        used_audio.add(edge["audio"])
        used_order.add(edge["order_number"])
    return selected


def build_corpus(
    assignments: list[dict[str, Any]],
    orders: list[dict[str, Any]],
    audios: list[dict[str, Any]],
    final_size: int,
) -> dict[str, Any]:
    order_by_number = {order["order_number"]: order for order in orders}
    audio_by_name = {audio["audio"]: audio for audio in audios}
    final_candidates = sorted(
        [edge for edge in assignments if edge["confidence_class"] == "metadata_exact"],
        key=lambda edge: edge["audio"],
        reverse=True,
    )
    if len(final_candidates) < final_size:
        raise RuntimeError(
            f"Seulement {len(final_candidates)} paires metadata_exact pour un holdout de {final_size}."
        )
    final_audio = {edge["audio"] for edge in final_candidates[:final_size]}
    rows: list[dict[str, Any]] = []
    for edge in sorted(assignments, key=lambda item: item["audio"]):
        order = order_by_number[edge["order_number"]]
        audio = audio_by_name[edge["audio"]]
        rows.append(
            {
                "audio": edge["audio"],
                "split": "final_holdout" if edge["audio"] in final_audio else "development",
                "transcription_sha256": audio["transcription_sha256"],
                "truth_order_number": order["order_number"],
                "truth_client": order["client_code"],
                "truth_delivery_date": order["delivery_date"],
                "truth_lines": order["lines"],
                "source": "independent_raw_transcript_pairer_v2",
                "pairing_confidence": {
                    key: edge[key]
                    for key in (
                        "confidence_class", "score", "client_score", "phone_unique",
                        "product_score", "product_coverage", "product_matches",
                        "audio_rank", "audio_margin", "order_rank", "order_margin",
                    )
                },
            }
        )
    return {
        "version": 2,
        "method": "independent_raw_transcript_and_metadata_mutual_matching",
        "matching_uses_program_product_predictions": False,
        "truth_visible_to_main_program": False,
        "split_policy": {
            "development": "paires restantes utilisables pour analyse",
            "final_holdout": f"{final_size} paires metadata_exact les plus recentes, interdites au developpement",
        },
        "rows": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--transcriptions", type=Path, required=True)
    parser.add_argument("--orders-csv", type=Path, action="append", required=True)
    parser.add_argument("--clients-xlsx", type=Path, required=True)
    parser.add_argument("--manual-phones", type=Path)
    parser.add_argument("--variants", type=Path)
    parser.add_argument("--date-from", required=True)
    parser.add_argument("--date-to", required=True)
    parser.add_argument("--final-size", type=int, default=20)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()

    clients, phones = load_clients(args.clients_xlsx, args.manual_phones, args.variants)
    audios = load_audios(args.transcriptions, args.date_from, args.date_to)
    orders = [
        order for order in load_orders(args.orders_csv)
        if args.date_from <= order["order_date"] <= args.date_to
    ]
    edges = build_edges(audios, orders, clients, phones)
    rank_edges(edges)
    classify_edges(edges, audios, orders, phones)
    assignments = select_assignments(edges)
    corpus = build_corpus(assignments, orders, audios, args.final_size)
    class_counts = Counter(edge["confidence_class"] for edge in assignments)
    split_counts = Counter(row["split"] for row in corpus["rows"])
    report = {
        "schema": "emalo-independent-pairing-report/v2",
        "audio_count": len(audios),
        "order_count": len(orders),
        "edge_count": len(edges),
        "assignment_count": len(assignments),
        "confidence_counts": dict(class_counts),
        "split_counts": dict(split_counts),
        "unmatched_audio_count": len(audios) - len(assignments),
        "matching_uses_program_product_predictions": False,
        "truth_visible_to_main_program": False,
        "copilote_send_attempted": False,
        "assignments": [
            {
                key: edge[key]
                for key in (
                    "audio", "order_number", "confidence_class", "score",
                    "client_score", "product_score", "product_coverage",
                    "audio_rank", "audio_margin", "order_rank", "order_margin",
                )
            }
            for edge in sorted(assignments, key=lambda item: item["audio"])
        ],
    }
    atomic_json(args.output, corpus)
    atomic_json(args.report, report)
    print(json.dumps({key: value for key, value in report.items() if key != "assignments"}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
