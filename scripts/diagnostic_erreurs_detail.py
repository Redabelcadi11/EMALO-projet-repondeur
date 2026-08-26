#!/usr/bin/env python3
"""
Diagnostic détaillé des articles manqués et faux positifs.
Lit les transcriptions réelles pour comprendre POURQUOI les articles sont ratés.
"""
from __future__ import annotations
import sys, os, json, re
sys.path.insert(0, str(__import__('pathlib').Path(__file__).resolve().parents[1]))
os.environ.setdefault("PYTHONDONTWRITEBYTECODE", "1")

from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "resultats" / "evaluation-copilote" / "comparaison_ES_vs_logiciel.xlsx"
TRANSCRIPTIONS = ROOT / "resultats" / "transcriptions"

# ── Charger catalogue pour désignations ──────────────────────────────────────
cat_path = ROOT / "config" / "catalogue-articles.json"
catalogue = {}
if cat_path.exists():
    for art in json.loads(cat_path.read_text("utf-8")):
        code = str(art.get("code_article") or "").strip().upper()
        if code:
            catalogue[code] = str(art.get("libelle_article") or "").strip()

# ── Charger le fichier comparaison ───────────────────────────────────────────
wb = load_workbook(EXCEL)
ws = wb[wb.sheetnames[1]]
headers = [cell.value for cell in ws[1]]
rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

# Regrouper par commande
commandes = []
current = None
for row in rows:
    if row.get("N°"):
        if current:
            commandes.append(current)
        current = {
            "num": row["N°"],
            "audio": row.get("Audio", ""),
            "articles": [],
        }
    if current:
        code = row.get("Code Article", "")
        if code:
            current["articles"].append({
                "code": code,
                "t_qty": row.get("Qté ES", ""),
                "t_unit": row.get("Unité ES", ""),
                "p_qty": row.get("Qté Logiciel", ""),
                "p_unit": row.get("Unité Logiciel", ""),
                "source": row.get("Source", ""),
            })
if current:
    commandes.append(current)

def get_transcription(audio_name: str) -> str:
    stem = Path(audio_name).stem
    # Remplacer extension .wav par __transcription.json
    t_path = TRANSCRIPTIONS / f"{stem}__transcription.json"
    if not t_path.exists():
        return ""
    try:
        data = json.loads(t_path.read_text("utf-8"))
        return str(data.get("texte") or data.get("transcription") or "").strip()
    except Exception:
        return ""

# ── Top articles manqués avec leurs transcriptions ───────────────────────────
from collections import Counter
manques_counter = Counter()
manques_par_code: dict[str, list[str]] = defaultdict(list)
fp_counter = Counter()
fp_par_code: dict[str, list[str]] = defaultdict(list)
qty_errors: list[dict] = []

for cmd in commandes:
    transcription = get_transcription(cmd["audio"])
    for a in cmd["articles"]:
        code = a["code"]
        if a["source"] == "ES seulement":
            manques_counter[code] += 1
            manques_par_code[code].append(transcription[:400])
        elif a["source"] == "Logiciel seulement":
            fp_counter[code] += 1
            fp_par_code[code].append(transcription[:400])
        elif a["source"] == "Les deux":
            if a["t_qty"] != a["p_qty"] or a["t_unit"] != a["p_unit"]:
                qty_errors.append({
                    "code": code,
                    "designation": catalogue.get(code, "???"),
                    "t_qty": a["t_qty"],
                    "t_unit": a["t_unit"],
                    "p_qty": a["p_qty"],
                    "p_unit": a["p_unit"],
                    "audio": cmd["audio"],
                    "transcription": transcription[:300],
                })

print("=" * 70)
print("ARTICLES MANQUES — transcriptions ou ils auraient du etre detectes")
print("=" * 70)
for code, cnt in manques_counter.most_common(10):
    lib = catalogue.get(code, "???")
    print(f"\n{'─'*60}")
    print(f"CODE={code} | {lib} | manqué {cnt}x")
    for i, texte in enumerate(manques_par_code[code][:2], 1):
        print(f"  TRANSCRIPTION {i}: {texte[:300]}")

print("\n" + "=" * 70)
print("FAUX POSITIFS — transcriptions ou ils ont ete proposes a tort")
print("=" * 70)
for code, cnt in fp_counter.most_common(8):
    lib = catalogue.get(code, "???")
    print(f"\n{'─'*60}")
    print(f"CODE={code} | {lib} | faux positif {cnt}x")
    for i, texte in enumerate(fp_par_code[code][:1], 1):
        print(f"  TRANSCRIPTION: {texte[:200]}")

print("\n" + "=" * 70)
print("ERREURS QUANTITE — analyse des patterns")
print("=" * 70)
for e in qty_errors[:20]:
    ratio = ""
    try:
        r = float(e["p_qty"]) / float(e["t_qty"])
        ratio = f" (ratio x{r:.1f})"
    except Exception:
        pass
    print(f"  {e['code']} ({e['designation'][:30]}): ES={e['t_qty']}{e['t_unit']} Log={e['p_qty']}{e['p_unit']}{ratio}")
    print(f"    -> {e['transcription'][:200]}")
