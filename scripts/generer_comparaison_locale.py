#!/usr/bin/env python3
"""
Comparaison ES vs Logiciel — Generateur Excel.

1. Apparier audios ↔ commandes ES (telephone → client + date audio = order_date).
2. Transcrire + analyser chaque audio via le worker distant.
3. Comparer ligne par ligne et generer un Excel avec coloration des differences.

AUCUN ENVOI A COPILOTE. Lecture seule.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

os.environ.setdefault("PYTHONDONTWRITEBYTECODE", "1")

from src.runtime_paths import bootstrap_runtime_environment

bootstrap_runtime_environment()

from extraire_informations import charger_clients
from src.clients import (
    charger_telephones_clients,
    enrichir_clients_avec_telephones,
    normaliser_telephone,
)
import worker_client

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
AUDIO_ROOT = ROOT / "ressources-originales" / "audio-nextcloud"
PHONE_CONFIG = ROOT / "config" / "telephones-clients.json"
TRUTH_CSV = ROOT / "resultats" / "copilote-replay" / "commandes_ES_200_dernieres_au_2026-08-11.csv"
OUTPUT_EXCEL = ROOT / "resultats" / "evaluation-copilote" / "comparaison_ES_vs_logiciel.xlsx"

AUDIO_DATE_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})_(?P<hour>\d{2})-(?P<minute>\d{2})-(?P<second>\d{2})"
)
PHONE_RE = re.compile(r"_De-([^_.]+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def audio_datetime(path: Path) -> datetime | None:
    match = AUDIO_DATE_RE.match(path.name)
    if not match:
        return None
    return datetime.strptime(
        f"{match.group('date')} {match.group('hour')}:{match.group('minute')}:{match.group('second')}",
        "%Y-%m-%d %H:%M:%S",
    )


def caller_phone(path: Path) -> str:
    match = PHONE_RE.search(path.name)
    return normaliser_telephone(match.group(1)) if match else ""


def unique_phone_clients() -> dict[str, str]:
    clients = charger_clients()
    enrichir_clients_avec_telephones(
        clients,
        charger_telephones_clients(PHONE_CONFIG),
    )
    codes_by_phone: dict[str, set[str]] = defaultdict(set)
    for client in clients:
        code = str(client.get("code_client") or "").strip()
        for value in client.get("telephones") or []:
            phone = normaliser_telephone(value)
            if phone and code:
                codes_by_phone[phone].add(code)
    return {
        phone: next(iter(codes))
        for phone, codes in codes_by_phone.items()
        if len(codes) == 1
    }


def date_iso(value: Any) -> str:
    match = re.search(r"\d{4}-\d{2}-\d{2}", str(value or ""))
    return match.group(0) if match else ""


def normalize_code(value: Any) -> str:
    text = str(value or "").strip()
    if "|" in text:
        text = text.split("|", 1)[0].strip()
    return text.upper()


def normalize_quantity(value: Any) -> str:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return ""
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text.upper()
    normalized = format(number.normalize(), "f")
    return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized


def normalize_unit(value: Any) -> str:
    return str(value or "").strip().upper()


def prediction_lines(command: dict[str, Any]) -> list[dict[str, str]]:
    return [
        {
            "code": normalize_code(line.get("code_article")),
            "designation": str(line.get("libelle_article") or ""),
            "quantity": normalize_quantity(line.get("quantite")),
            "unit": normalize_unit(line.get("unite")),
        }
        for line in command.get("lignes_commande") or []
        if normalize_code(line.get("code_article"))
    ]


# ---------------------------------------------------------------------------
# Read truth CSV
# ---------------------------------------------------------------------------


def read_truth(path: Path) -> dict[str, dict[str, Any]]:
    orders: dict[str, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter=";"):
            number = str(row.get("order_number") or "").strip()
            if not number:
                continue
            order = orders.setdefault(
                number,
                {
                    "order_number": number,
                    "client_code": str(row.get("client_code") or "").strip(),
                    "client_label": str(row.get("client_label") or "").strip(),
                    "delivery_date": date_iso(row.get("delivery_date")),
                    "order_date": date_iso(row.get("order_date")),
                    "departure_date": date_iso(row.get("departure_date")),
                    "lines": [],
                    "errors": [],
                },
            )
            error = str(row.get("error") or "").strip()
            if error:
                order["errors"].append(error)
            article = normalize_code(row.get("article_code"))
            designation = str(row.get("designation") or "").strip()
            if article:
                order["lines"].append(
                    {
                        "code": article,
                        "designation": designation,
                        "quantity": normalize_quantity(row.get("quantity")),
                        "unit": normalize_unit(row.get("unit")),
                    }
                )
    return {
        number: order
        for number, order in orders.items()
        if not order["errors"] and order["lines"]
    }


# ---------------------------------------------------------------------------
# Improved pairing: phone → client + audio_date == order_date + 1-to-1
# ---------------------------------------------------------------------------


def select_pairs(
    truth: dict[str, dict[str, Any]],
    audio_start: date,
    audio_end: date,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    phone_clients = unique_phone_clients()

    # Index audios by (client_code, audio_date)
    AudioEntry = dict[str, Any]
    audios_by_client_date: dict[tuple[str, str], list[Path]] = defaultdict(list)
    scanned = 0
    mapped = 0
    for path in sorted(AUDIO_ROOT.glob("*")):
        when = audio_datetime(path)
        if when is None:
            continue
        audio_date = when.date()
        if audio_date < audio_start or audio_date > audio_end:
            continue
        scanned += 1
        code = phone_clients.get(caller_phone(path), "")
        if code:
            mapped += 1
            audios_by_client_date[(code, audio_date.isoformat())].append(path)

    # Index orders by (client_code, order_date)
    orders_by_client_date: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for order in truth.values():
        code = str(order.get("client_code") or "").strip()
        order_date = order.get("order_date", "")
        if code and order_date:
            orders_by_client_date[(code, order_date)].append(order)

    # Relaxed pairing: For each client, match audio with order if within ±2 days
    from datetime import timedelta
    pairs = []
    paired_orders = set()
    paired_audios = set()
    
    for client, audios in audios_by_client_date.items():
        if client[0] not in [c for c, d in orders_by_client_date.keys()]:
            continue
            
        # Get all orders for this client
        client_orders = []
        for (c, d), ords in orders_by_client_date.items():
            if c == client[0]:
                for o in ords:
                    if o["order_number"] not in paired_orders:
                        client_orders.append((date.fromisoformat(d), o))
                        
        for audio in audios:
            if audio.name in paired_audios:
                continue
            
            a_date = date.fromisoformat(client[1])
            # Find closest order within ±2 days
            best_order = None
            best_diff = 999
            for o_date, o in client_orders:
                diff = abs((o_date - a_date).days)
                if diff <= 2 and diff < best_diff:
                    best_diff = diff
                    best_order = o
            
            if best_order:
                pairs.append({
                    "audio": audio,
                    "truth": best_order,
                    "phone_client_code": client[0],
                    "match_type": "fuzzy_date",
                })
                paired_orders.add(best_order["order_number"])
                paired_audios.add(audio.name)
                # Remove from client_orders so it doesn't get matched twice
                client_orders = [(d, o) for d, o in client_orders if o["order_number"] != best_order["order_number"]]

    # Fallback: try matching on client only (across all dates in window) for remaining
    remaining_orders_by_client: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for order in truth.values():
        if order["order_number"] in paired_orders:
            continue
        code = str(order.get("client_code") or "").strip()
        if code:
            remaining_orders_by_client[code].append(order)

    remaining_audios_by_client: dict[str, list[Path]] = defaultdict(list)
    for (code, _date), audio_list in audios_by_client_date.items():
        for audio in audio_list:
            if audio.name not in paired_audios:
                remaining_audios_by_client[code].append(audio)

    for code in sorted(set(remaining_audios_by_client) & set(remaining_orders_by_client)):
        audios = remaining_audios_by_client[code]
        orders = remaining_orders_by_client[code]
        if len(audios) == 1 and len(orders) == 1:
            pairs.append({
                "audio": audios[0],
                "truth": orders[0],
                "phone_client_code": code,
                "match_type": "client_only",
            })

    return pairs, {
        "audios_scanned": scanned,
        "audios_with_unambiguous_phone": mapped,
        "truth_orders": len(truth),
        "strict_pairs": len([p for p in pairs if p["match_type"] == "strict_date"]),
        "client_only_pairs": len([p for p in pairs if p["match_type"] == "client_only"]),
        "total_pairs": len(pairs),
    }


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def generate_excel(pairs: list[dict[str, Any]], coverage: dict[str, int]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ----- Feuille Resume -----
    ws_resume = wb.active
    ws_resume.title = "Résumé"
    ws_resume.append(["Métrique", "Valeur"])
    for key, val in coverage.items():
        ws_resume.append([key, val])

    # ----- Feuille Comparaison -----
    ws = wb.create_sheet("Comparaison détaillée")

    headers = [
        "N°",
        "Audio",
        "Appariement",
        "N° Commande ES",
        "Client ES (code)",
        "Client ES (nom)",
        "Client Logiciel",
        "Client OK ?",
        "Date Liv ES",
        "Date Liv Logiciel",
        "Date OK ?",
        "Code Article",
        "Désignation",
        "Qté ES",
        "Unité ES",
        "Qté Logiciel",
        "Unité Logiciel",
        "Qté OK ?",
        "Unité OK ?",
        "Source",
        "Statut Logiciel",
    ]
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    missing_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Stats
    total = len(pairs)
    perfect_count = 0
    client_ok_count = 0
    date_ok_count = 0
    total_truth_lines = 0
    total_pred_lines = 0
    code_match_count = 0

    for pair_idx, pair in enumerate(pairs, 1):
        audio = pair["audio"]
        t_order = pair["truth"]
        command = pair.get("predicted_command", {})
        match_type = pair.get("match_type", "")

        t_client = str(t_order["client_code"])
        t_client_label = str(t_order.get("client_label", ""))
        p_client = str(command.get("client_retenu") or "")

        t_date = t_order["delivery_date"]
        delivery = command.get("date_livraison", {})
        p_date = date_iso(delivery.get("date_iso") if isinstance(delivery, dict) else delivery)

        status = command.get("statut", "")

        # Lines
        t_lines = t_order.get("lines", [])
        p_lines = prediction_lines(command)

        total_truth_lines += len(t_lines)
        total_pred_lines += len(p_lines)

        client_ok = t_client == p_client
        date_ok = t_date == p_date
        if client_ok:
            client_ok_count += 1
        if date_ok:
            date_ok_count += 1

        # Group by code for aligned display
        t_by_code: dict[str, list[dict]] = defaultdict(list)
        p_by_code: dict[str, list[dict]] = defaultdict(list)
        for line in t_lines:
            t_by_code[line["code"]].append(line)
        for line in p_lines:
            p_by_code[line["code"]].append(line)

        all_codes = sorted(set(t_by_code.keys()) | set(p_by_code.keys()))
        code_matches_local = 0
        for code in all_codes:
            t_count = len(t_by_code.get(code, []))
            p_count = len(p_by_code.get(code, []))
            code_matches_local += min(t_count, p_count)
        code_match_count += code_matches_local

        # Build aligned rows
        aligned_rows: list[dict] = []
        for code in all_codes:
            t_list = t_by_code.get(code, [])
            p_list = p_by_code.get(code, [])
            max_count = max(len(t_list), len(p_list))
            for i in range(max_count):
                t_l = t_list[i] if i < len(t_list) else {}
                p_l = p_list[i] if i < len(p_list) else {}
                source = ""
                if t_l and p_l:
                    source = "Les deux"
                elif t_l:
                    source = "ES seulement"
                elif p_l:
                    source = "Logiciel seulement"
                aligned_rows.append({
                    "code": code,
                    "designation": t_l.get("designation") or p_l.get("designation", ""),
                    "t_qty": t_l.get("quantity", ""),
                    "t_unit": t_l.get("unit", ""),
                    "p_qty": p_l.get("quantity", ""),
                    "p_unit": p_l.get("unit", ""),
                    "source": source,
                })

        # Check perfect
        t_exact = Counter((l["code"], l["quantity"], l["unit"]) for l in t_lines)
        p_exact = Counter((l["code"], l["quantity"], l["unit"]) for l in p_lines)
        perfect = client_ok and date_ok and t_exact == p_exact
        if perfect:
            perfect_count += 1

        if not aligned_rows:
            aligned_rows = [{"code": "", "designation": "", "t_qty": "", "t_unit": "", "p_qty": "", "p_unit": "", "source": ""}]

        for i, arow in enumerate(aligned_rows):
            row = [
                pair_idx if i == 0 else "",
                audio.name if i == 0 else "",
                match_type if i == 0 else "",
                t_order["order_number"] if i == 0 else "",
                t_client if i == 0 else "",
                t_client_label if i == 0 else "",
                p_client if i == 0 else "",
                "✓" if client_ok else "✗" if i == 0 else "",
                t_date if i == 0 else "",
                p_date if i == 0 else "",
                "✓" if date_ok else "✗" if i == 0 else "",
                arow["code"],
                arow["designation"],
                arow["t_qty"],
                arow["t_unit"],
                arow["p_qty"],
                arow["p_unit"],
                "✓" if arow["t_qty"] == arow["p_qty"] and arow["t_qty"] else ("✗" if arow["t_qty"] or arow["p_qty"] else ""),
                "✓" if arow["t_unit"] == arow["p_unit"] and arow["t_unit"] else ("✗" if arow["t_unit"] or arow["p_unit"] else ""),
                arow["source"],
                status if i == 0 else "",
            ]
            ws.append(row)
            current_row = ws.max_row

            # Stripe alternate pairs
            if pair_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    if not cell.fill or cell.fill.start_color.rgb in ("00000000", None):
                        cell.fill = stripe_fill

            # Color client column
            if i == 0:
                ws.cell(row=current_row, column=7).fill = ok_fill if client_ok else err_fill
                ws.cell(row=current_row, column=8).fill = ok_fill if client_ok else err_fill
                ws.cell(row=current_row, column=10).fill = ok_fill if date_ok else err_fill
                ws.cell(row=current_row, column=11).fill = ok_fill if date_ok else err_fill

            # Color article source
            if arow["source"] == "ES seulement":
                ws.cell(row=current_row, column=12).fill = err_fill
                ws.cell(row=current_row, column=20).fill = err_fill
            elif arow["source"] == "Logiciel seulement":
                ws.cell(row=current_row, column=12).fill = warn_fill
                ws.cell(row=current_row, column=20).fill = warn_fill

            # Color quantity/unit
            if arow["t_qty"] or arow["p_qty"]:
                qty_ok = arow["t_qty"] == arow["p_qty"]
                ws.cell(row=current_row, column=16).fill = ok_fill if qty_ok else err_fill
                ws.cell(row=current_row, column=18).fill = ok_fill if qty_ok else err_fill
            if arow["t_unit"] or arow["p_unit"]:
                unit_ok = arow["t_unit"] == arow["p_unit"]
                ws.cell(row=current_row, column=17).fill = ok_fill if unit_ok else err_fill
                ws.cell(row=current_row, column=19).fill = ok_fill if unit_ok else err_fill

    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Add stats to resume
    ws_resume.append([])
    ws_resume.append(["Statistiques de comparaison", ""])
    ws_resume.append(["Total paires appariées", total])
    ws_resume.append(["Commandes parfaites", perfect_count])
    ws_resume.append(["Taux commandes parfaites", f"{perfect_count / total * 100:.1f}%" if total else "N/A"])
    ws_resume.append(["Client correct", client_ok_count])
    ws_resume.append(["Taux client correct", f"{client_ok_count / total * 100:.1f}%" if total else "N/A"])
    ws_resume.append(["Date livraison correcte", date_ok_count])
    ws_resume.append(["Taux date correcte", f"{date_ok_count / total * 100:.1f}%" if total else "N/A"])
    ws_resume.append(["Total lignes ES (vérité)", total_truth_lines])
    ws_resume.append(["Total lignes Logiciel", total_pred_lines])
    ws_resume.append(["Correspondances code article", code_match_count])
    recall = code_match_count / total_truth_lines * 100 if total_truth_lines else 0
    precision = code_match_count / total_pred_lines * 100 if total_pred_lines else 0
    ws_resume.append(["Rappel produit (code)", f"{recall:.1f}%"])
    ws_resume.append(["Précision produit (code)", f"{precision:.1f}%"])

    # Style resume
    for cell in ws_resume[1]:
        cell.font = Font(bold=True)

    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_EXCEL)
    return OUTPUT_EXCEL


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    print("=" * 70)
    print("COMPARAISON ES vs LOGICIEL — Lecture seule, zéro envoi Copilote")
    print("=" * 70)

    # 1. Read truth
    print("\n1. Lecture vérité ES (Tous les fichiers)...")
    import glob
    truth = {}
    csv_pattern = str(ROOT / "resultats" / "copilote-replay" / "*.csv")
    for csv_file in glob.glob(csv_pattern):
        if "commandes_ES_" in csv_file or "historique_es_" in csv_file:
            try:
                t = read_truth(Path(csv_file))
                truth.update(t)
            except:
                pass
    print(f"   {len(truth)} commandes ES valides trouvées")

    # 2. Pair audios
    print("\n2. Appariement audios <-> commandes ES...")
    pairs, coverage = select_pairs(truth, date(2026, 4, 1), date(2026, 8, 30))
    print(f"   Audios scannés : {coverage['audios_scanned']}")
    print(f"   Audios avec téléphone reconnu : {coverage['audios_with_unambiguous_phone']}")
    print(f"   Paires strictes (client+date) : {coverage['strict_pairs']}")
    print(f"   Paires client seul : {coverage['client_only_pairs']}")
    print(f"   TOTAL PAIRES : {coverage['total_pairs']}")

    if not pairs:
        print("\nAucune paire trouvée. Arrêt.")
        return 1

    # 3. Analyze each audio via remote worker
    print(f"\n3. Analyse via worker distant ({len(pairs)} audios)...")
    print("   (transcription Whisper large-v3 + extraction métier)")

    for idx, pair in enumerate(pairs, 1):
        audio = pair["audio"]
        print(f"   [{idx}/{len(pairs)}] {audio.name}...", end=" ", flush=True)
        started = time.perf_counter()
        try:
            result = worker_client.remote_analyze_audio(audio)
            elapsed = time.perf_counter() - started
            if not result.get("ok"):
                print(f"ERREUR: {result.get('message', '?')} ({elapsed:.1f}s)")
                pair["predicted_command"] = {
                    "client_retenu": "",
                    "date_livraison": {},
                    "lignes_commande": [],
                    "statut": "ERREUR_WORKER",
                }
                continue

            worker_client.write_remote_transcription(audio, result)
            commands = result.get("commandes")
            if isinstance(commands, list) and len(commands) >= 1:
                pair["predicted_command"] = commands[0]
            else:
                pair["predicted_command"] = {
                    "client_retenu": "",
                    "date_livraison": {},
                    "lignes_commande": [],
                    "statut": "PROBLEMATIQUE",
                }
            print(f"OK ({elapsed:.1f}s)")
        except Exception as exc:
            elapsed = time.perf_counter() - started
            print(f"EXCEPTION: {exc} ({elapsed:.1f}s)")
            pair["predicted_command"] = {
                "client_retenu": "",
                "date_livraison": {},
                "lignes_commande": [],
                "statut": "ERREUR",
            }

    # 4. Generate Excel
    print("\n4. Génération du fichier Excel...")
    output = generate_excel(pairs, coverage)
    print(f"   Fichier : {output}")
    print("\nTerminé !")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
