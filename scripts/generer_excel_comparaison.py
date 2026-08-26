import json
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json-report", required=True)
    parser.add_argument("--output-excel", required=True)
    args = parser.parse_args()

    json_path = Path(args.json_report)
    if not json_path.exists():
        print(f"File {json_path} does not exist.")
        return 1

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison ES vs Logiciel"

    # Define headers
    headers = [
        "Audio",
        "Vérité Client (ES)",
        "Logiciel Client",
        "Vérité Date Liv",
        "Logiciel Date Liv",
        "Article",
        "Vérité Qte",
        "Vérité Unité",
        "Logiciel Qte",
        "Logiciel Unité",
        "Statut Logiciel",
        "Parfaite ?"
    ]
    
    ws.append(headers)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for order in data.get("orders", []):
        audio = order.get("audio")
        t_client = order.get("truth_client")
        p_client = order.get("predicted_client")
        t_date = order.get("truth_delivery_date")
        p_date = order.get("predicted_delivery_date")
        status = order.get("program_status")
        perfect = "Oui" if order.get("perfect_order") else "Non"

        # Align lines by code
        t_lines = {L["code"]: L for L in order.get("truth_lines", [])}
        p_lines = {L["code"]: L for L in order.get("predicted_lines", [])}
        
        all_codes = sorted(set(t_lines.keys()) | set(p_lines.keys()))
        
        if not all_codes:
            # If no lines for some reason
            ws.append([audio, t_client, p_client, t_date, p_date, "", "", "", "", "", status, perfect])
            continue
            
        for i, code in enumerate(all_codes):
            t_L = t_lines.get(code, {})
            p_L = p_lines.get(code, {})
            
            # Print common order info only on the first line for readability, or on all lines
            row = [
                audio if i == 0 else "",
                t_client if i == 0 else "",
                p_client if i == 0 else "",
                t_date if i == 0 else "",
                p_date if i == 0 else "",
                code,
                t_L.get("quantity", ""),
                t_L.get("unit", ""),
                p_L.get("quantity", ""),
                p_L.get("unit", ""),
                status if i == 0 else "",
                perfect if i == 0 else ""
            ]
            ws.append(row)
            
            # Highlight differences
            current_row = ws.max_row
            
            if i == 0:
                if t_client != p_client:
                    ws.cell(row=current_row, column=3).fill = err_fill
                else:
                    ws.cell(row=current_row, column=3).fill = ok_fill
                    
                if t_date != p_date:
                    ws.cell(row=current_row, column=5).fill = err_fill
                else:
                    ws.cell(row=current_row, column=5).fill = ok_fill

            # Check quantity/unit match
            if t_L.get("quantity", "") != p_L.get("quantity", "") or t_L.get("unit", "") != p_L.get("unit", ""):
                ws.cell(row=current_row, column=9).fill = err_fill
                ws.cell(row=current_row, column=10).fill = err_fill
            else:
                ws.cell(row=current_row, column=9).fill = ok_fill
                ws.cell(row=current_row, column=10).fill = ok_fill

    output_path = Path(args.output_excel)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel generated at: {output_path.resolve()}")

if __name__ == '__main__':
    main()
