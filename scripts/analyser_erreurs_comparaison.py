#!/usr/bin/env python3
"""Analyse les patterns d'erreurs dans le fichier de comparaison."""
import sys
import os
sys.path.insert(0, str(__import__('pathlib').Path(__file__).resolve().parents[1]))
os.environ.setdefault("PYTHONDONTWRITEBYTECODE", "1")

from openpyxl import load_workbook
from collections import defaultdict, Counter
from pathlib import Path

EXCEL = Path(__file__).resolve().parents[1] / "resultats" / "evaluation-copilote" / "comparaison_ES_vs_logiciel.xlsx"

wb = load_workbook(EXCEL)
print("Feuilles:", wb.sheetnames)
ws = wb[wb.sheetnames[1]]  # Comparaison détaillée

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(headers, row))
    rows.append(r)

print(f"Total lignes (articles): {len(rows)}")

# Regrouper par commande (N° non vide = première ligne d'une commande)
commandes = []
current = None
for row in rows:
    if row.get("N°"):  # Premiere ligne d'une nouvelle commande
        if current:
            commandes.append(current)
        current = {
            "num": row["N°"],
            "audio": row.get("Audio", ""),
            "match_type": row.get("Appariement", ""),
            "n_commande_es": row.get("N° Commande ES", ""),
            "client_es": row.get("Client ES (code)", ""),
            "client_logiciel": row.get("Client Logiciel", ""),
            "client_ok": row.get("Client OK ?", ""),
            "date_es": row.get("Date Liv ES", ""),
            "date_logiciel": row.get("Date Liv Logiciel", ""),
            "date_ok": row.get("Date OK ?", ""),
            "statut": row.get("Statut Logiciel", ""),
            "articles": [],
        }
    if current:
        code = row.get("Code Article", "")
        if code:
            current["articles"].append({
                "code": code,
                "designation": row.get("Désignation", ""),
                "t_qty": row.get("Qté ES", ""),
                "t_unit": row.get("Unité ES", ""),
                "p_qty": row.get("Qté Logiciel", ""),
                "p_unit": row.get("Unité Logiciel", ""),
                "source": row.get("Source", ""),
            })
if current:
    commandes.append(current)

print(f"Total commandes: {len(commandes)}")

# ---------------------------------------------------------------
# Analyses
# ---------------------------------------------------------------

# 1. Dates
date_errors = [c for c in commandes if c["date_ok"] == "✗"]
date_ok = [c for c in commandes if c["date_ok"] == "✓"]
print(f"\n=== DATES ===")
print(f"Date OK: {len(date_ok)} | Date ERREUR: {len(date_errors)}")
for c in date_errors:
    print(f"  ES={c['date_es']} | Logiciel={c['date_logiciel']} | audio={c['audio'][:30]}")

# 2. Articles manquants (ES seulement = logiciel n'a pas trouvé)
print(f"\n=== ARTICLES MANQUES (ES seulement - logiciel ne les a pas trouvés) ===")
manques_counter = Counter()
for c in commandes:
    for a in c["articles"]:
        if a["source"] == "ES seulement":
            manques_counter[a["code"]] += 1

for code, cnt in manques_counter.most_common(20):
    print(f"  {code}: manqué {cnt}x")

# 3. Articles en trop (Logiciel seulement = faux positifs)
print(f"\n=== FAUX POSITIFS (Logiciel seulement - pas dans commande ES) ===")
fp_counter = Counter()
for c in commandes:
    for a in c["articles"]:
        if a["source"] == "Logiciel seulement":
            fp_counter[a["code"]] += 1
for code, cnt in fp_counter.most_common(20):
    print(f"  {code}: faux positif {cnt}x")

# 4. Erreurs quantité/unité (article trouvé mais mauvaise qte/unite)
print(f"\n=== ERREURS QUANTITE/UNITE (article trouvé mais mauvaise qte ou unite) ===")
unit_errors = []
qty_errors = []
for c in commandes:
    for a in c["articles"]:
        if a["source"] == "Les deux":
            if a["t_qty"] != a["p_qty"]:
                qty_errors.append(a)
                print(f"  QTE  code={a['code']} ES={a['t_qty']} Log={a['p_qty']}")
            if a["t_unit"] != a["p_unit"]:
                unit_errors.append(a)
                print(f"  UNIT code={a['code']} ES={a['t_unit']} Log={a['p_unit']}")

# 5. Statuts logiciel
print(f"\n=== STATUTS LOGICIEL ===")
statuts = Counter(c["statut"] for c in commandes)
for s, n in statuts.most_common():
    print(f"  {s}: {n}")

# 6. Stats globales articles
es_only = sum(1 for c in commandes for a in c["articles"] if a["source"] == "ES seulement")
log_only = sum(1 for c in commandes for a in c["articles"] if a["source"] == "Logiciel seulement")
les_deux = sum(1 for c in commandes for a in c["articles"] if a["source"] == "Les deux")
print(f"\n=== RESUME ARTICLES ===")
print(f"  Trouvés par les deux : {les_deux}")
print(f"  Manqués (ES seulement) : {es_only}")
print(f"  Faux positifs (Logiciel seul) : {log_only}")
print(f"  Erreurs quantité : {len(qty_errors)}")
print(f"  Erreurs unité : {len(unit_errors)}")
