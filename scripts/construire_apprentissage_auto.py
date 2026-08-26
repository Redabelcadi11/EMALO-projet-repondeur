import json
import re
from pathlib import Path
from openpyxl import load_workbook
import string

try:
    from rapidfuzz import fuzz
except ImportError:
    import sys
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "rapidfuzz"])
    from rapidfuzz import fuzz

EXCEL = Path(__file__).resolve().parents[1] / "resultats" / "evaluation-copilote" / "comparaison_ES_vs_logiciel.xlsx"
TRANSCRIPTIONS_DIR = Path(__file__).resolve().parents[1] / "resultats" / "transcriptions"
REGLES_PATH = Path(__file__).resolve().parents[1] / "config" / "regles-apprentissage.json"

STOPWORDS = {"et", "le", "la", "les", "un", "une", "des", "du", "de", "pour", "demain", "bonjour", "au", "aux", "il", "faut", "je", "voudrais", "svp", "merci", "kilo", "kilos", "litre", "litres", "carton", "cartons"}

def normalize(text):
    text = str(text).lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return " ".join(text.split())

def main():
    wb = load_workbook(EXCEL)
    ws = wb[wb.sheetnames[1]]
    headers = [cell.value for cell in ws[1]]
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        rows.append(r)

    try:
        with open(REGLES_PATH, "r", encoding="utf-8") as f:
            regles_data = json.load(f)
    except Exception:
        regles_data = {"version": 1, "rules": []}
    
    regles_data["rules"] = [r for r in regles_data.get("rules", []) if r.get("origin") != "auto_apprentissage"]
    existing_rules = regles_data["rules"]
    
    current_order = None
    new_rules = []

    for row in rows:
        if row.get("N°"):
            current_order = row
            audio_name = row.get("Audio", "")
            json_name = audio_name.replace(".wav", "__transcription.json")
            json_path = TRANSCRIPTIONS_DIR / json_name
            transcription_text = ""
            if json_path.exists():
                with open(json_path, "r", encoding="utf-8") as f:
                    tdata = json.load(f)
                    transcription_text = normalize(tdata.get("texte", ""))
            current_order["texte"] = transcription_text

        if current_order and current_order.get("texte"):
            qte_es_str = str(row.get("Qté ES", "")).strip()
            qte_log_str = str(row.get("Qté Logiciel", "")).strip()
            
            qte_es = float(qte_es_str) if qte_es_str.replace('.','',1).isdigit() else 0
            qte_log = float(qte_log_str) if qte_log_str.replace('.','',1).isdigit() else 0
            
            if qte_es > 0 and qte_log == 0:
                designation = normalize(row.get("Désignation", ""))
                client = str(current_order.get("Client ES (code)", "")).strip()
                code_article = str(row.get("Code Article", "")).strip()
                
                words = [w for w in current_order["texte"].split() if w not in STOPWORDS and len(w) > 2]
                best_phrase = ""
                best_score = 0
                
                for i in range(len(words)):
                    for j in range(i+1, min(i+4, len(words)+1)):
                        phrase = " ".join(words[i:j])
                        score_token = fuzz.token_set_ratio(designation, phrase)
                        
                        if score_token > best_score:
                            best_score = score_token
                            best_phrase = phrase
                
                # If we found a good fuzzy match (score > 80), it means the client said it differently
                if best_score > 80 and len(best_phrase) > 4:
                    rule_id = f"auto_{client}_{code_article}_{best_phrase.replace(' ', '_')}"
                    if not any(r.get("id") == rule_id for r in existing_rules + new_rules):
                        new_rules.append({
                            "id": rule_id,
                            "client_all": [client],
                            "mention_any": [best_phrase],
                            "label_all": [row.get("Désignation", "")],
                            "bonus": 90,
                            "enabled": True,
                            "origin": "auto_apprentissage"
                        })
            
    regles_data["rules"].extend(new_rules)
    
    with open(REGLES_PATH, "w", encoding="utf-8") as f:
        json.dump(regles_data, f, indent=2, ensure_ascii=False)
    
    print(f"Added {len(new_rules)} new rules from historical comparison.")

if __name__ == "__main__":
    main()
